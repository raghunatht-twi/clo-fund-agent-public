"""Reads each data product sheet of the workbook into list[dict].

Supports two funds via a sheet-name prefix map:
  DKIG-2024-VII  →  sheets "DP-01 Static Profile", "DP-02 Portfolio Snapshot", …
  DKIG-2016-I    →  sheets "2016-DP-01 Static Profile", "2016-DP-02 Portfolio Snapshot", …
"""

from __future__ import annotations

import hashlib
import os
from functools import lru_cache
from typing import Any

from openpyxl import load_workbook

from . import WORKBOOK_PATH

HEADER_ROW = 4   # rows 1-3 are title/subtitle/blank

FUND_SHEET_PREFIX: dict[str, str] = {
    "DKIG-2024-VII": "",
    "DKIG-2016-I":   "2016-",
}
VALID_FUNDS = frozenset(FUND_SHEET_PREFIX)

# ---------------------------------------------------------------------------
# AI-06: Workbook integrity check
# ---------------------------------------------------------------------------
_workbook_hash: str | None = None


def _verify_workbook() -> None:
    """Verify workbook SHA-256 on first load and guard against runtime tampering.

    If CLO_WORKBOOK_SHA256 is set, the file must match exactly.
    Without the env var, the hash is recorded at first access and compared on
    every subsequent call to detect modification during a live session.
    """
    global _workbook_hash
    with open(WORKBOOK_PATH, "rb") as f:
        current_hash = hashlib.sha256(f.read()).hexdigest()
    expected = os.environ.get("CLO_WORKBOOK_SHA256", "")
    if expected and current_hash != expected:
        raise EnvironmentError(
            f"Workbook integrity check failed "
            f"(expected …{expected[-12:]}, got …{current_hash[-12:]})."
        )
    if _workbook_hash is None:
        _workbook_hash = current_hash
    elif current_hash != _workbook_hash:
        raise EnvironmentError(
            "Workbook file changed on disk since startup — possible tampering."
        )


def _read_sheet(sheet_name: str, fund_id: str = "DKIG-2024-VII") -> list[dict[str, Any]]:
    if fund_id not in VALID_FUNDS:
        raise ValueError(f"Unknown fund_id {fund_id!r}. Valid values: {sorted(VALID_FUNDS)}")
    _verify_workbook()
    prefix = FUND_SHEET_PREFIX[fund_id]
    wb = load_workbook(WORKBOOK_PATH, data_only=True, read_only=True)
    ws = wb[prefix + sheet_name]
    rows = list(ws.iter_rows(min_row=HEADER_ROW, values_only=True))
    if not rows:
        return []
    headers = [h for h in rows[0] if h is not None]
    width = len(headers)
    data: list[dict[str, Any]] = []
    for raw in rows[1:]:
        if raw is None or all(v is None for v in raw):
            continue
        record = {headers[i]: raw[i] for i in range(width)}
        data.append(record)
    wb.close()
    return data


@lru_cache(maxsize=4)
def static_profile(fund_id: str = "DKIG-2024-VII") -> dict[str, Any]:
    """DP-01 stored as Attribute/Value pairs — flatten to a single dict."""
    raw = _read_sheet("DP-01 Static Profile", fund_id)
    return {row["Attribute"]: row["Value"] for row in raw}


@lru_cache(maxsize=4)
def portfolio(fund_id: str = "DKIG-2024-VII") -> list[dict[str, Any]]:
    return _read_sheet("DP-02 Portfolio Snapshot", fund_id)


@lru_cache(maxsize=4)
def performance(fund_id: str = "DKIG-2024-VII") -> list[dict[str, Any]]:
    return _read_sheet("DP-03 Performance", fund_id)


@lru_cache(maxsize=4)
def compliance(fund_id: str = "DKIG-2024-VII") -> list[dict[str, Any]]:
    return _read_sheet("DP-04 Compliance", fund_id)


@lru_cache(maxsize=4)
def cashflows(fund_id: str = "DKIG-2024-VII") -> list[dict[str, Any]]:
    return _read_sheet("DP-05 Cashflows", fund_id)


@lru_cache(maxsize=4)
def fees(fund_id: str = "DKIG-2024-VII") -> list[dict[str, Any]]:
    return _read_sheet("DP-06 Fees", fund_id)


@lru_cache(maxsize=4)
def key_metrics(fund_id: str = "DKIG-2024-VII") -> list[dict[str, Any]]:
    return _read_sheet("DP-07 Key Metrics", fund_id)


@lru_cache(maxsize=4)
def liability_structure(fund_id: str = "DKIG-2024-VII") -> list[dict[str, Any]]:
    return _read_sheet("DP-08 Liability Structure", fund_id)
