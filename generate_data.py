"""
Generates a brand-styled Excel workbook of synthetic CLO fund data
covering data products DP-01 through DP-08, plus a JSON-LD ontology
that an agent can load to interpret and query the workbook.

Single fund modelled: DKIG Funding 2024-VII LLC, reporting 2026-04-30.
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
XLSX_PATH = ROOT / "CLO_Fund_Domain_Data.xlsx"
ONT_PATH = ROOT / "clo-fund-ontology.jsonld"

# ---------------------------------------------------------------------------
# Brand styling (per CLAUDE.md)
# ---------------------------------------------------------------------------
COLORS = {
    "white": "FFFFFF",
    "panel": "EDF1F3",
    "ink": "000000",
    "primary": "003D4F",
    "coral": "F2617A",
    "amber": "CC850A",
    "green": "689E78",
    "teal": "47A1AD",
    "purple": "634F7D",
}

TITLE_FONT = Font(name="Bitter", size=18, bold=True, color=COLORS["primary"])
SUBTITLE_FONT = Font(name="Inter", size=11, italic=True, color=COLORS["primary"])
HEADER_FONT = Font(name="Inter", size=10, bold=True, color=COLORS["white"])
BODY_FONT = Font(name="Inter", size=10, color=COLORS["ink"])
BAND_FONT = Font(name="Inter", size=10, color=COLORS["ink"])

HEADER_FILL = PatternFill("solid", fgColor=COLORS["primary"])
BAND_FILL = PatternFill("solid", fgColor=COLORS["panel"])
PASS_FILL = PatternFill("solid", fgColor=COLORS["green"])
FAIL_FILL = PatternFill("solid", fgColor=COLORS["coral"])

THIN = Side(style="thin", color=COLORS["panel"])
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_sheet(
    wb,
    sheet_name: str,
    dp_id: str,
    title: str,
    subtitle: str,
    columns: list[str],
    rows: list[tuple],
    pass_col: int | None = None,
    number_formats: dict[str, str] | None = None,
) -> None:
    """Write one data-product sheet with consistent branded styling."""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    # Title block
    ws.cell(row=1, column=1, value=f"{dp_id}  ·  {title}").font = TITLE_FONT
    ws.cell(row=2, column=1, value=subtitle).font = SUBTITLE_FONT
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=len(columns))
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=len(columns))
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18

    # Header
    header_row = 4
    for c, label in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=c, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[header_row].height = 24

    # Data rows
    number_formats = number_formats or {}
    for r, row in enumerate(rows, start=header_row + 1):
        is_band = (r - header_row) % 2 == 0
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = BAND_FONT if is_band else BODY_FONT
            if is_band:
                cell.fill = BAND_FILL
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = BORDER
            col_label = columns[c - 1]
            if col_label in number_formats:
                cell.number_format = number_formats[col_label]
            # Pass/Fail colour highlight
            if pass_col is not None and c - 1 == pass_col:
                if str(value).upper() == "PASS":
                    cell.fill = PASS_FILL
                    cell.font = Font(name="Inter", size=10, bold=True, color=COLORS["white"])
                elif str(value).upper() == "FAIL":
                    cell.fill = FAIL_FILL
                    cell.font = Font(name="Inter", size=10, bold=True, color=COLORS["white"])

    # Column widths
    for c, label in enumerate(columns, start=1):
        max_len = max([len(str(label))] + [len(str(r[c - 1])) for r in rows] or [10])
        ws.column_dimensions[get_column_letter(c)].width = min(max(max_len + 2, 12), 36)


# ---------------------------------------------------------------------------
# Reference data
# ---------------------------------------------------------------------------
FUND_ID = "DKIG-2024-VII"
REPORTING_DATE = "2026-04-30"

# ---------------------------------------------------------------------------
# DP-01 Fund Static Profile
# ---------------------------------------------------------------------------
DP01_COLS = ["Attribute", "Value"]
DP01_ROWS = [
    ("Fund ID", FUND_ID),
    ("Fund Name", "DKIG Funding 2024-VII LLC"),
    ("CUSIP / ISIN", "12529TAA0 / US12529TAA08"),
    ("Manager Name", "DKIG Asset Management LLC"),
    ("Trustee", "U.S. Bank Trust Company, N.A."),
    ("Vintage Year", 2024),
    ("Closing Date", "2024-03-15"),
    ("Reinvestment Period End", "2029-03-15"),
    ("Non-Call Period End", "2026-03-15"),
    ("Legal Final Maturity", "2037-03-15"),
    ("Target Collateral Type", "Broadly Syndicated 1st Lien Senior Secured Loans"),
    ("Governing Law", "State of New York"),
    ("Base Currency", "USD"),
    ("Senior Management Fee Rate (% p.a.)", 0.0015),
    ("Subordinated Management Fee Rate (% p.a.)", 0.0025),
    ("Incentive Fee Hurdle (IRR %)", 0.12),
    ("Incentive Fee Catch-up", 0.20),
    ("Target Par Amount (USD)", 500_000_000),
]
DP01_FORMATS = {"Value": "#,##0.00####"}

# ---------------------------------------------------------------------------
# DP-08 Fund Liability Structure (defined early — used by other sheets)
# ---------------------------------------------------------------------------
DP08_COLS = [
    "Tranche Class", "CUSIP", "Initial Notional (USD)", "Current Notional (USD)",
    "Coupon Type", "Coupon Rate (SOFR+ bps)", "Payment Frequency",
    "Moody's Rating", "S&P Rating", "Fitch Rating", "Subordination Level (%)",
    "OC Cushion (%)", "IC Cushion (%)", "Waterfall Priority",
    "Cumulative Principal Repaid (USD)", "Interest Paid Current Period (USD)",
    "Interest Accrued (USD)", "Rating Outlook",
]
DP08_ROWS = [
    ("A", "12529TAB8", 310_000_000, 310_000_000, "Floating", 148, "Quarterly",
     "Aaa", "AAA", "AAA", 38.0, 3.8, 4.6, 1, 0, 1_211_500, 403_833, "Stable"),
    ("B", "12529TAC6", 50_000_000, 50_000_000, "Floating", 200, "Quarterly",
     "Aa2", "AA", "AA", 28.0, 4.8, 4.2, 2, 0, 230_000, 76_667, "Stable"),
    ("C", "12529TAD4", 30_000_000, 30_000_000, "Floating", 260, "Quarterly",
     "A2", "A", "A", 22.0, 4.8, 3.6, 3, 0, 165_000, 55_000, "Stable"),
    ("D", "12529TAE2", 25_000_000, 25_000_000, "Floating", 335, "Quarterly",
     "Baa3", "BBB-", "BBB-", 17.0, 4.9, 2.7, 4, 0, 168_750, 56_250, "Stable"),
    ("E", "12529TAF9", 20_000_000, 20_000_000, "Floating", 650, "Quarterly",
     "Ba3", "BB-", "BB-", 13.0, 4.2, 1.8, 5, 0, 195_000, 65_000, "Negative"),
    ("Sub Notes", "12529TAG7", 30_000_000, 30_000_000, "Residual", 0, "Quarterly",
     "NR", "NR", "NR", 7.0, None, None, 6, 0, 0, 0, "NR"),
    ("Equity", "12529TAH5", 35_000_000, 35_000_000, "Residual", 0, "Quarterly",
     "NR", "NR", "NR", 0.0, None, None, 7, 0, 0, 0, "NR"),
]
DP08_FORMATS = {
    "Initial Notional (USD)": "$#,##0",
    "Current Notional (USD)": "$#,##0",
    "Subordination Level (%)": "0.00",
    "OC Cushion (%)": "0.00",
    "IC Cushion (%)": "0.00",
    "Cumulative Principal Repaid (USD)": "$#,##0",
    "Interest Paid Current Period (USD)": "$#,##0",
    "Interest Accrued (USD)": "$#,##0",
    "Coupon Rate (SOFR+ bps)": "0",
    "Waterfall Priority": "0",
}

# ---------------------------------------------------------------------------
# DP-02 Fund Portfolio Snapshot — 22 sample loan positions
# ---------------------------------------------------------------------------
DP02_COLS = [
    "Position ID", "Obligor Name", "Facility CUSIP", "Industry (Moody's)",
    "Country", "Loan Type", "Par Amount (USD)", "Market Value (USD)",
    "Price (% par)", "Spread (SOFR+ bps)", "Maturity Date",
    "Moody's Rating", "S&P Rating", "Fitch Rating", "PIK Flag",
    "LBO Flag", "Covenant-Lite Flag", "Days Past Due",
]
DP02_ROWS = [
    ("P0001", "Asurion LLC", "04685TAJ7", "Services: Business", "USA", "1st Lien", 12_000_000, 11_820_000, 98.50, 325, "2031-08-19", "B2", "B", "B", "N", "Y", "Y", 0),
    ("P0002", "Athenahealth Inc.", "047111AC2", "Healthcare & Pharma", "USA", "1st Lien", 9_500_000, 9_281_500, 97.70, 325, "2029-02-15", "B2", "B-", "B", "N", "Y", "Y", 0),
    ("P0003", "Boxer Parent Co Inc (BMC)", "10316PAA8", "High Tech Industries", "USA", "1st Lien", 10_000_000, 9_950_000, 99.50, 400, "2028-10-02", "B2", "B-", "B", "N", "Y", "Y", 0),
    ("P0004", "Caesars Entertainment Inc.", "12769GAA0", "Hotel, Gaming & Leisure", "USA", "1st Lien", 8_000_000, 7_968_000, 99.60, 275, "2030-02-06", "B1", "B+", "BB-", "N", "N", "Y", 0),
    ("P0005", "Charter Communications", "16117MAU8", "Media: Broadcasting", "USA", "1st Lien", 14_000_000, 13_846_000, 98.90, 200, "2027-04-30", "Ba1", "BB+", "BB+", "N", "N", "Y", 0),
    ("P0006", "Clarios Global LP", "18068KAA1", "Automotive", "USA", "1st Lien", 7_500_000, 7_357_500, 98.10, 325, "2030-05-06", "B1", "B+", "BB-", "N", "Y", "Y", 0),
    ("P0007", "DaVita Inc.", "23918KAR4", "Healthcare & Pharma", "USA", "1st Lien", 6_000_000, 5_982_000, 99.70, 175, "2028-08-12", "Ba2", "BB", "BB", "N", "N", "Y", 0),
    ("P0008", "Endo International plc", "29273RBE8", "Healthcare & Pharma", "USA", "1st Lien", 5_500_000, 4_785_000, 87.00, 425, "2028-04-23", "Caa1", "CCC+", "CCC", "N", "N", "Y", 0),
    ("P0009", "Frontier Communications", "35906AAB1", "Telecommunications", "USA", "1st Lien", 7_000_000, 6_790_000, 97.00, 375, "2031-05-01", "B1", "BB-", "BB-", "N", "N", "Y", 0),
    ("P0010", "Hilton Worldwide Holdings", "43300AAH7", "Hotel, Gaming & Leisure", "USA", "1st Lien", 9_000_000, 8_982_000, 99.80, 175, "2028-06-22", "Ba1", "BB+", "BBB-", "N", "N", "Y", 0),
    ("P0011", "Iron Mountain Inc.", "46284VAH3", "Services: Business", "USA", "1st Lien", 6_500_000, 6_435_000, 99.00, 225, "2029-01-31", "Ba3", "BB-", "BB", "N", "N", "Y", 0),
    ("P0012", "Level 3 Financing", "52729NBJ0", "Telecommunications", "USA", "1st Lien", 8_000_000, 7_360_000, 92.00, 350, "2029-03-15", "B2", "B+", "B+", "N", "N", "Y", 0),
    ("P0013", "Medline Industries", "58404DAL8", "Healthcare & Pharma", "USA", "1st Lien", 11_500_000, 11_385_000, 99.00, 275, "2028-10-23", "B1", "B+", "BB-", "N", "Y", "Y", 0),
    ("P0014", "Mileage Plus Holdings (United)", "59556DAA1", "Transportation: Cargo", "USA", "1st Lien", 5_000_000, 5_000_000, 100.00, 525, "2027-06-21", "Ba1", "BB+", "BB+", "N", "N", "Y", 0),
    ("P0015", "Numericable / Altice France", "67054KAA8", "Telecommunications", "FRA", "1st Lien", 6_000_000, 5_400_000, 90.00, 525, "2028-08-15", "Caa1", "B-", "B", "N", "Y", "Y", 0),
    ("P0016", "PetSmart LLC", "71678KAB2", "Retail", "USA", "1st Lien", 7_500_000, 7_237_500, 96.50, 375, "2028-02-11", "B1", "B", "B+", "N", "Y", "Y", 0),
    ("P0017", "Restaurant Brands Intl.", "76131DAB6", "Beverage, Food & Tobacco", "CAN", "1st Lien", 8_500_000, 8_457_500, 99.50, 175, "2030-09-21", "Ba2", "BB+", "BB+", "N", "N", "Y", 0),
    ("P0018", "Scientific Games Holdings", "80874PAR4", "Hotel, Gaming & Leisure", "USA", "1st Lien", 7_000_000, 6_790_000, 97.00, 350, "2029-04-04", "B2", "B", "B", "N", "Y", "Y", 0),
    ("P0019", "Tenneco Inc.", "88032QAA0", "Automotive", "USA", "1st Lien", 6_500_000, 6_175_000, 95.00, 425, "2028-11-17", "B2", "B", "B", "N", "Y", "Y", 0),
    ("P0020", "Univision Communications", "91342NAB5", "Media: Broadcasting", "USA", "1st Lien", 7_000_000, 6_650_000, 95.00, 425, "2029-06-24", "B2", "B", "B+", "N", "Y", "Y", 0),
    ("P0021", "Verscend Holding (Cotiviti)", "92535MAA6", "High Tech Industries", "USA", "1st Lien", 6_000_000, 5_910_000, 98.50, 325, "2028-08-27", "B2", "B-", "B", "N", "Y", "Y", 0),
    ("P0022", "Zayo Group Holdings", "98919JAA1", "Telecommunications", "USA", "1st Lien", 8_500_000, 7_905_000, 93.00, 425, "2027-03-09", "B3", "CCC+", "B-", "Y", "Y", "Y", 0),
]
DP02_FORMATS = {
    "Par Amount (USD)": "$#,##0",
    "Market Value (USD)": "$#,##0",
    "Price (% par)": "0.00",
    "Spread (SOFR+ bps)": "0",
    "Days Past Due": "0",
}

# ---------------------------------------------------------------------------
# DP-03 Fund Performance Metrics — last 12 monthly snapshots
# ---------------------------------------------------------------------------
DP03_COLS = [
    "Reporting Date", "Total Fund NAV (USD)", "Equity NAV (USD)",
    "Gross IRR (%)", "Net IRR (%)", "DPI", "RVPI", "TVPI",
    "Inception-to-Date P&L (USD)", "Current Period P&L (USD)",
    "Unrealised G/L (USD)", "Realised G/L (USD)",
    "Total Interest Income (USD)", "Benchmark Return (%)",
    "Excess Return vs Benchmark (%)",
]
DP03_ROWS = [
    ("2025-05-31", 498_750_000, 33_750_000, 0.108, 0.092, 0.05, 1.06, 1.11, 1_350_000, 510_000, 800_000, 50_000, 21_460_000, 0.082, 0.026),
    ("2025-06-30", 499_125_000, 34_125_000, 0.111, 0.094, 0.05, 1.08, 1.13, 1_980_000, 630_000, 1_120_000, 60_000, 24_210_000, 0.083, 0.028),
    ("2025-07-31", 499_500_000, 34_500_000, 0.114, 0.096, 0.06, 1.09, 1.15, 2_640_000, 660_000, 1_500_000, 80_000, 26_950_000, 0.084, 0.030),
    ("2025-08-31", 499_725_000, 34_725_000, 0.116, 0.098, 0.07, 1.10, 1.17, 3_270_000, 630_000, 1_690_000, 90_000, 29_700_000, 0.085, 0.031),
    ("2025-09-30", 499_950_000, 34_950_000, 0.118, 0.100, 0.07, 1.11, 1.18, 3_900_000, 630_000, 1_870_000, 110_000, 32_450_000, 0.086, 0.032),
    ("2025-10-31", 500_175_000, 35_175_000, 0.120, 0.102, 0.08, 1.12, 1.20, 4_545_000, 645_000, 2_065_000, 130_000, 35_200_000, 0.087, 0.033),
    ("2025-11-30", 500_400_000, 35_400_000, 0.121, 0.103, 0.09, 1.13, 1.22, 5_175_000, 630_000, 2_245_000, 150_000, 37_950_000, 0.088, 0.033),
    ("2025-12-31", 500_625_000, 35_625_000, 0.123, 0.104, 0.10, 1.14, 1.24, 5_820_000, 645_000, 2_430_000, 170_000, 40_700_000, 0.089, 0.034),
    ("2026-01-31", 500_775_000, 35_775_000, 0.124, 0.105, 0.10, 1.15, 1.25, 6_435_000, 615_000, 2_590_000, 200_000, 43_400_000, 0.090, 0.034),
    ("2026-02-28", 500_900_000, 35_900_000, 0.125, 0.106, 0.11, 1.15, 1.26, 7_050_000, 615_000, 2_750_000, 220_000, 46_080_000, 0.091, 0.034),
    ("2026-03-31", 501_050_000, 36_050_000, 0.126, 0.107, 0.11, 1.16, 1.27, 7_650_000, 600_000, 2_905_000, 250_000, 48_750_000, 0.092, 0.035),
    ("2026-04-30", 501_225_000, 36_225_000, 0.127, 0.108, 0.12, 1.17, 1.29, 8_265_000, 615_000, 3_080_000, 270_000, 51_410_000, 0.093, 0.035),
]
DP03_FORMATS = {
    "Total Fund NAV (USD)": "$#,##0",
    "Equity NAV (USD)": "$#,##0",
    "Gross IRR (%)": "0.00%",
    "Net IRR (%)": "0.00%",
    "DPI": "0.00",
    "RVPI": "0.00",
    "TVPI": "0.00",
    "Inception-to-Date P&L (USD)": "$#,##0",
    "Current Period P&L (USD)": "$#,##0",
    "Unrealised G/L (USD)": "$#,##0",
    "Realised G/L (USD)": "$#,##0",
    "Total Interest Income (USD)": "$#,##0",
    "Benchmark Return (%)": "0.00%",
    "Excess Return vs Benchmark (%)": "0.00%",
}

# ---------------------------------------------------------------------------
# DP-04 Fund Compliance Dashboard
# ---------------------------------------------------------------------------
DP04_COLS = [
    "Test ID", "Test Name", "Test Type", "Tranche Class",
    "Current Value", "Threshold", "Cushion", "Pass/Fail",
    "Breach Consequence", "Last Tested",
]
DP04_ROWS = [
    ("OC-A",  "Class A/B OC Test",   "OC",        "A/B",  1.283, 1.245, 0.038, "PASS",
     "Divert principal & interest proceeds to repay Class A until cured",   REPORTING_DATE),
    ("OC-C",  "Class C OC Test",     "OC",        "C",    1.218, 1.170, 0.048, "PASS",
     "Divert proceeds to repay Class A then Class B principal",            REPORTING_DATE),
    ("OC-D",  "Class D OC Test",     "OC",        "D",    1.154, 1.105, 0.049, "PASS",
     "Divert proceeds to repay senior tranches in priority order",         REPORTING_DATE),
    ("OC-E",  "Class E OC Test",     "OC",        "E",    1.097, 1.055, 0.042, "PASS",
     "Divert proceeds to repay senior tranches in priority order",         REPORTING_DATE),
    ("IC-A",  "Class A/B IC Test",   "IC",        "A/B",  1.460, 1.200, 0.046, "PASS",
     "Divert proceeds to repay Class A principal",                         REPORTING_DATE),
    ("IC-C",  "Class C IC Test",     "IC",        "C",    1.336, 1.150, 0.036, "PASS",
     "Divert proceeds to repay senior tranches",                           REPORTING_DATE),
    ("IC-D",  "Class D IC Test",     "IC",        "D",    1.227, 1.105, 0.027, "PASS",
     "Divert proceeds to repay senior tranches",                           REPORTING_DATE),
    ("IC-E",  "Class E IC Test",     "IC",        "E",    1.118, 1.085, 0.018, "PASS",
     "Divert proceeds to repay senior tranches",                           REPORTING_DATE),
    ("WARF",  "Moody's WARF Covenant", "Quality", "Fund", 2715,  2900,  185,   "PASS",
     "Trading restrictions; CCC bucket reclassification",                  REPORTING_DATE),
    ("DIV",   "Diversity Score",     "Quality",   "Fund", 86.0,  80.0,  6.0,   "PASS",
     "Trading restrictions; new purchase ineligibility",                   REPORTING_DATE),
    ("CCC",   "CCC/Caa Bucket %",    "Concentration", "Fund", 0.062, 0.075, 0.013, "PASS",
     "Excess CCC haircut to market value in OC test",                      REPORTING_DATE),
    ("OBL",   "Largest Single Obligor %", "Concentration", "Fund", 0.029, 0.030, 0.001, "PASS",
     "Trading restrictions until cured",                                   REPORTING_DATE),
    ("IND",   "Largest Single Industry %", "Concentration", "Fund", 0.107, 0.150, 0.043, "PASS",
     "Trading restrictions until cured",                                   REPORTING_DATE),
    ("PIK",   "DIP/PIK %",           "Concentration", "Fund", 0.017, 0.075, 0.058, "PASS",
     "Excess treated as defaulted in OC test",                             REPORTING_DATE),
]
DP04_FORMATS = {
    "Current Value": "0.0000",
    "Threshold": "0.0000",
    "Cushion": "0.0000",
}

# ---------------------------------------------------------------------------
# DP-05 Fund Cashflow Statement — last 4 quarterly payment dates
# ---------------------------------------------------------------------------
DP05_COLS = [
    "Payment Date", "Collection Period", "Total Interest Proceeds (USD)",
    "Total Principal Proceeds (USD)", "Reinvestment Proceeds (USD)",
    "Recoveries (USD)", "Waterfall Step", "Recipient",
    "Amount Disbursed (USD)", "OC Diversion Amount (USD)",
    "Equity Distribution Amount (USD)", "Management Fee Paid (USD)",
    "Incentive Fee Paid (USD)", "Trustee Fee Paid (USD)",
]


def waterfall_period(
    payment_date: str,
    collection_period: str,
    interest: float,
    principal: float,
    reinvestment: float,
    equity_dist: float,
    mgmt_fee: float,
    incentive_fee: float,
    trustee_fee: float,
) -> list[tuple]:
    """Generate full 9-step waterfall rows for a single payment period."""
    rows = []
    base = (payment_date, collection_period, interest, principal, reinvestment, 0)
    rows.append((*base, "1. Senior Expenses", "Trustee + Admin + Tax", trustee_fee + 75_000, 0, 0, 0, 0, trustee_fee))
    rows.append((*base, "2. Class A Interest", "Class A Tranche", 1_211_500, 0, 0, 0, 0, 0))
    rows.append((*base, "3. Class A OC Test", "Test passes — no diversion", 0, 0, 0, 0, 0, 0))
    rows.append((*base, "4. Class A IC Test", "Test passes — no diversion", 0, 0, 0, 0, 0, 0))
    rows.append((*base, "5. Class B–E Interest & Tests", "Class B/C/D/E Tranches", 230_000 + 165_000 + 168_750 + 195_000, 0, 0, 0, 0, 0))
    rows.append((*base, "6. Senior Management Fee", "DKIG Asset Management", mgmt_fee * 0.375, 0, 0, mgmt_fee * 0.375, 0, 0))
    rows.append((*base, "7. Subordinated Notes Interest", "Sub Notes Holders", 525_000, 0, 0, 0, 0, 0))
    rows.append((*base, "8. Incentive Fee / Sub Mgmt Fee", "DKIG Asset Management", incentive_fee + mgmt_fee * 0.625, 0, 0, mgmt_fee * 0.625, incentive_fee, 0))
    rows.append((*base, "9. Equity Distribution", "Preference Shareholders (Equity)", equity_dist, 0, equity_dist, 0, 0, 0))
    return rows


DP05_ROWS = []
DP05_ROWS += waterfall_period("2025-07-21", "2025-04-21 → 2025-07-20", 8_750_000, 0, 0,
                              equity_dist=4_410_750, mgmt_fee=500_000, incentive_fee=0, trustee_fee=45_000)
DP05_ROWS += waterfall_period("2025-10-21", "2025-07-21 → 2025-10-20", 9_120_000, 0, 0,
                              equity_dist=4_710_750, mgmt_fee=500_000, incentive_fee=0, trustee_fee=45_000)
DP05_ROWS += waterfall_period("2026-01-21", "2025-10-21 → 2026-01-20", 9_540_000, 0, 0,
                              equity_dist=5_010_750, mgmt_fee=500_000, incentive_fee=0, trustee_fee=45_000)
DP05_ROWS += waterfall_period("2026-04-21", "2026-01-21 → 2026-04-20", 9_785_000, 0, 0,
                              equity_dist=5_215_750, mgmt_fee=500_000, incentive_fee=0, trustee_fee=45_000)

DP05_FORMATS = {
    "Total Interest Proceeds (USD)": "$#,##0",
    "Total Principal Proceeds (USD)": "$#,##0",
    "Reinvestment Proceeds (USD)": "$#,##0",
    "Recoveries (USD)": "$#,##0",
    "Amount Disbursed (USD)": "$#,##0",
    "OC Diversion Amount (USD)": "$#,##0",
    "Equity Distribution Amount (USD)": "$#,##0",
    "Management Fee Paid (USD)": "$#,##0",
    "Incentive Fee Paid (USD)": "$#,##0",
    "Trustee Fee Paid (USD)": "$#,##0",
}

# ---------------------------------------------------------------------------
# DP-06 Fund Fee & Expense Ledger
# ---------------------------------------------------------------------------
DP06_COLS = [
    "Period", "Fee Type", "Fee Rate / Amount", "Accrued YTD (USD)",
    "Accrued Current Period (USD)", "Amount Paid Current Period (USD)",
    "Cumulative Amount Paid (USD)", "Hurdle Rate (%)", "Catch-up (%)",
    "Tax Provision (USD)", "Effective Tax Rate (%)", "Total Expense Ratio (%)",
]
DP06_ROWS = [
    ("2026 Q1", "Management Fee — Senior",       "0.15% p.a.",         562_500,  187_500,  187_500, 2_625_000, None,  None,  None,    None,   None),
    ("2026 Q1", "Management Fee — Subordinated", "0.25% p.a.",         937_500,  312_500,  312_500, 4_375_000, None,  None,  None,    None,   None),
    ("2026 Q1", "Incentive Fee",                 "20% above 12% IRR",        0,        0,        0,         0, 0.12,  0.20,  None,    None,   None),
    ("2026 Q1", "Trustee Fee",                   "USD 180,000 p.a.",    45_000,   45_000,   45_000,   315_000, None,  None,  None,    None,   None),
    ("2026 Q1", "Admin / Accounting Fee",        "USD 100,000 p.a.",    25_000,   25_000,   25_000,   175_000, None,  None,  None,    None,   None),
    ("2026 Q1", "Legal Fee",                     "Variable",            45_000,   45_000,   30_000,   210_000, None,  None,  None,    None,   None),
    ("2026 Q1", "Rating Agency Fee",             "USD 200,000 p.a.",    50_000,   50_000,   50_000,   350_000, None,  None,  None,    None,   None),
    ("2026 Q1", "Tax Provision",                 "Effective",                0,        0,        0,         0, None,  None,  410_000, 0.0210, None),
    ("2026 YTD", "Total Expense Ratio (TER)",    "Aggregate",       1_665_000,  665_000,  650_000, 8_050_000, None,  None,  410_000, 0.0210, 0.0094),
]
DP06_FORMATS = {
    "Accrued YTD (USD)": "$#,##0",
    "Accrued Current Period (USD)": "$#,##0",
    "Amount Paid Current Period (USD)": "$#,##0",
    "Cumulative Amount Paid (USD)": "$#,##0",
    "Hurdle Rate (%)": "0.00%",
    "Catch-up (%)": "0.00%",
    "Tax Provision (USD)": "$#,##0",
    "Effective Tax Rate (%)": "0.00%",
    "Total Expense Ratio (%)": "0.00%",
}

# ---------------------------------------------------------------------------
# DP-07 Fund Key Metrics Tracker — last 8 weeks
# ---------------------------------------------------------------------------
DP07_COLS = [
    "Reporting Date", "WAS (bps over SOFR)", "WARF", "WAL (years)", "WAC (%)",
    "Weighted Avg Recovery Rate (%)", "Par Build/Loss vs Target (USD)",
    "% Floating Rate", "% Fixed Rate", "% PIK", "% CCC/Caa", "% Covenant-Lite",
    "Diversity Score", "Number of Obligors", "Number of Industries",
    "Largest Single Obligor (%)", "Largest Single Industry (%)",
    "Top 10 Obligor Concentration (%)",
]
DP07_ROWS = [
    ("2026-03-09", 318, 2705, 5.7, 0.082, 0.62, 1_125_000, 1.00, 0.00, 0.014, 0.061, 0.78, 85.0, 168, 28, 0.030, 0.108, 0.244),
    ("2026-03-16", 320, 2710, 5.6, 0.083, 0.62, 1_240_000, 1.00, 0.00, 0.015, 0.061, 0.78, 85.5, 168, 28, 0.030, 0.108, 0.245),
    ("2026-03-23", 321, 2712, 5.6, 0.083, 0.62, 1_310_000, 1.00, 0.00, 0.016, 0.062, 0.78, 85.5, 167, 28, 0.029, 0.108, 0.246),
    ("2026-03-30", 322, 2714, 5.5, 0.084, 0.62, 1_400_000, 1.00, 0.00, 0.016, 0.062, 0.78, 86.0, 167, 28, 0.029, 0.107, 0.246),
    ("2026-04-06", 323, 2715, 5.5, 0.084, 0.62, 1_460_000, 1.00, 0.00, 0.017, 0.062, 0.78, 86.0, 167, 28, 0.029, 0.107, 0.247),
    ("2026-04-13", 323, 2716, 5.5, 0.084, 0.62, 1_510_000, 1.00, 0.00, 0.017, 0.062, 0.78, 86.0, 167, 28, 0.029, 0.107, 0.247),
    ("2026-04-20", 324, 2715, 5.4, 0.085, 0.62, 1_570_000, 1.00, 0.00, 0.017, 0.062, 0.78, 86.0, 167, 28, 0.029, 0.107, 0.248),
    ("2026-04-30", 325, 2715, 5.4, 0.085, 0.62, 1_625_000, 1.00, 0.00, 0.017, 0.062, 0.78, 86.0, 167, 28, 0.029, 0.107, 0.248),
]
DP07_FORMATS = {
    "WAS (bps over SOFR)": "0",
    "WARF": "0",
    "WAL (years)": "0.0",
    "WAC (%)": "0.00%",
    "Weighted Avg Recovery Rate (%)": "0.00%",
    "Par Build/Loss vs Target (USD)": "$#,##0",
    "% Floating Rate": "0.00%",
    "% Fixed Rate": "0.00%",
    "% PIK": "0.00%",
    "% CCC/Caa": "0.00%",
    "% Covenant-Lite": "0.00%",
    "Diversity Score": "0.0",
    "Number of Obligors": "0",
    "Number of Industries": "0",
    "Largest Single Obligor (%)": "0.00%",
    "Largest Single Industry (%)": "0.00%",
    "Top 10 Obligor Concentration (%)": "0.00%",
}

# ============================================================================
# FUND 2 — DKIG Funding 2016-I LLC  (2016 Vintage, now amortising)
# ============================================================================
FUND2_ID = "DKIG-2016-I"
FUND2_REPORTING_DATE = "2026-03-31"
_F2_EQ_INITIAL = 44_000_000  # initial equity investment — denominator for DPI/RVPI

def _build_performance_row(
    reporting_date: str,
    nav: float,
    eq_nav: float,
    g_irr: float,
    n_irr: float,
    dpi: float,
    rvpi: float,
    itd_pl: float,
    qtr_pl: float,
    unrealised: float,
    realised: float,
    interest: float,
    bench: float,
    excess: float,
) -> tuple:
    """Build one DP-03 row, computing TVPI = DPI + RVPI."""
    return (reporting_date, nav, eq_nav, g_irr, n_irr, dpi, rvpi,
            round(dpi + rvpi, 3),
            itd_pl, qtr_pl, unrealised, realised, interest, bench, excess)

# DP-01 — Fund Static Profile
F2_DP01_ROWS = [
    ("Fund ID",                              FUND2_ID),
    ("Fund Name",                            "DKIG Funding 2016-I LLC"),
    ("CUSIP / ISIN",                         "12529BAA1 / US12529BAA14"),
    ("Manager Name",                         "DKIG Asset Management LLC"),
    ("Trustee",                              "U.S. Bank Trust Company, N.A."),
    ("Vintage Year",                         2016),
    ("Closing Date",                         "2016-04-15"),
    ("Reinvestment Period End",              "2020-04-15"),
    ("Non-Call Period End",                  "2018-04-15"),
    ("Legal Final Maturity",                 "2029-04-15"),
    ("Target Collateral Type",               "Broadly Syndicated 1st Lien Senior Secured Loans"),
    ("Governing Law",                        "State of New York"),
    ("Base Currency",                        "USD"),
    ("Senior Management Fee Rate (% p.a.)",  0.0015),
    ("Subordinated Management Fee Rate (% p.a.)", 0.0025),
    ("Incentive Fee Hurdle (IRR %)",         0.12),
    ("Incentive Fee Catch-up",               0.20),
    ("Target Par Amount (USD)",              400_000_000),
]

# DP-08 — Liability Structure (Class A mostly amortised after 10 yrs)
F2_DP08_ROWS = [
    ("A",        "12529BAB9", 240_000_000,  72_000_000, "Floating", 130, "Quarterly",
     "Aaa", "AAA", "AAA", 18.0, 5.2, 6.1, 1, 168_000_000,   810_000, 270_000, "Stable"),
    ("B",        "12529BAC7",  38_000_000,  30_000_000, "Floating", 185, "Quarterly",
     "Aa2",  "AA",  "AA",  7.5, 5.8, 5.4, 2,   8_000_000,   138_750,  46_250, "Stable"),
    ("C",        "12529BAD5",  22_000_000,  19_500_000, "Floating", 245, "Quarterly",
      "A2",   "A",   "A",  5.0, 4.5, 4.1, 3,   2_500_000,   119_438,  39_813, "Stable"),
    ("D",        "12529BAE3",  18_000_000,  16_800_000, "Floating", 320, "Quarterly",
     "Baa3", "BBB-","BBB-", 3.8, 3.2, 3.0, 4,   1_200_000,   134_400,  44_800, "Stable"),
    ("E",        "12529BAF0",  16_000_000,  15_200_000, "Floating", 620, "Quarterly",
      "Ba3",  "BB-", "BB-", 3.0, 2.8, 2.1, 5,     800_000,   235_600,  78_533, "Negative"),
    ("Sub Notes","12529BAG8",  22_000_000,  22_000_000, "Residual",   0, "Quarterly",
       "NR",   "NR",  "NR", 0.0, None, None, 6,         0,         0,       0, "NR"),
    ("Equity",   "12529BAH6",  44_000_000,  44_000_000, "Residual",   0, "Quarterly",
       "NR",   "NR",  "NR", 0.0, None, None, 7,         0,         0,       0, "NR"),
]

# DP-02 — Portfolio Snapshot (current, ~$175M par remaining, 15 positions)
F2_DP02_ROWS = [
    ("Q0001","Asurion LLC",            "04685TAJ7","Services: Business",    "USA","1st Lien", 14_000_000, 13_580_000, 97.00, 350,"2030-08-19","B2","B","B",  "N","Y","Y",0),
    ("Q0002","Athenahealth Inc.",       "047111AC2","Healthcare & Pharma",   "USA","1st Lien", 11_000_000, 10_670_000, 97.00, 325,"2029-02-15","B2","B-","B", "N","Y","Y",0),
    ("Q0003","Charter Communications", "16117MAU8","Media: Broadcasting",   "USA","1st Lien", 13_500_000, 13_365_000, 99.00, 200,"2027-04-30","Ba1","BB+","BB+","N","N","Y",0),
    ("Q0004","DaVita Inc.",             "23918KAR4","Healthcare & Pharma",   "USA","1st Lien",  9_000_000,  8_955_000, 99.50, 175,"2028-08-12","Ba2","BB","BB","N","N","Y",0),
    ("Q0005","Frontier Communications", "35906AAB1","Telecommunications",    "USA","1st Lien", 10_000_000,  9_700_000, 97.00, 375,"2031-05-01","B1","BB-","BB-","N","N","Y",0),
    ("Q0006","Hilton Worldwide",        "43300AAH7","Hotel, Gaming & Leisure","USA","1st Lien", 12_000_000, 11_988_000, 99.90, 175,"2028-06-22","Ba1","BB+","BBB-","N","N","Y",0),
    ("Q0007","Iron Mountain Inc.",      "46284VAH3","Services: Business",    "USA","1st Lien",  8_000_000,  7_920_000, 99.00, 225,"2029-01-31","Ba3","BB-","BB","N","N","Y",0),
    ("Q0008","Medline Industries",      "58404DAL8","Healthcare & Pharma",   "USA","1st Lien", 14_000_000, 13_860_000, 99.00, 275,"2028-10-23","B1","B+","BB-","N","Y","Y",0),
    ("Q0009","PetSmart LLC",            "71678KAB2","Retail",                "USA","1st Lien",  9_500_000,  9_215_000, 97.00, 375,"2028-02-11","B1","B","B+","N","Y","Y",0),
    ("Q0010","Restaurant Brands Intl.","76131DAB6","Beverage, Food & Tobacco","CAN","1st Lien",11_000_000, 10_945_000, 99.50, 175,"2030-09-21","Ba2","BB+","BB+","N","N","Y",0),
    ("Q0011","Scientific Games",        "80874PAR4","Hotel, Gaming & Leisure","USA","1st Lien",  9_000_000,  8_730_000, 97.00, 350,"2029-04-04","B2","B","B","N","Y","Y",0),
    ("Q0012","Tenneco Inc.",            "88032QAA0","Automotive",            "USA","1st Lien",  8_500_000,  7_990_000, 94.00, 450,"2028-11-17","Caa1","CCC+","B-","N","Y","Y",0),
    ("Q0013","Univision Communications","91342NAB5","Media: Broadcasting",   "USA","1st Lien",  9_000_000,  8_550_000, 95.00, 425,"2029-06-24","B2","B","B+","N","Y","Y",0),
    ("Q0014","Verscend Holding",        "92535MAA6","High Tech Industries",  "USA","1st Lien", 10_000_000,  9_850_000, 98.50, 325,"2028-08-27","B2","B-","B","N","Y","Y",0),
    ("Q0015","Zayo Group Holdings",     "98919JAA1","Telecommunications",    "USA","1st Lien", 10_000_000,  9_200_000, 92.00, 450,"2027-03-09","B3","CCC+","B-","Y","Y","Y",0),
]

# DP-03 — Performance: 40 quarters Q2-2016 → Q1-2026
F2_DP03_ROWS = [
    # 2016 — Ramp-up (3 qtrs)
    _build_performance_row("2016-06-30", 398_500_000, 44_200_000, 0.022, 0.000, 0.00, 1.005,    200_000,    200_000,    150_000,    50_000,  3_920_000, 0.062, -0.040),
    _build_performance_row("2016-09-30", 400_200_000, 44_700_000, 0.048, 0.025, 0.00, 1.016,    700_000,    500_000,    420_000,    80_000,  4_050_000, 0.063, -0.015),
    _build_performance_row("2016-12-31", 402_100_000, 45_400_000, 0.070, 0.048, 0.02, 1.031,  1_400_000,    700_000,  1_000_000,   200_000,  4_150_000, 0.064,  0.006),
    # 2017 — Building momentum (4 qtrs)
    _build_performance_row("2017-03-31", 403_500_000, 46_000_000, 0.082, 0.061, 0.03, 1.045,  2_000_000,    600_000,  1_400_000,   250_000,  4_350_000, 0.066,  0.016),
    _build_performance_row("2017-06-30", 405_000_000, 46_600_000, 0.090, 0.069, 0.04, 1.059,  2_600_000,    600_000,  1_800_000,   280_000,  4_480_000, 0.067,  0.023),
    _build_performance_row("2017-09-30", 406_500_000, 47_200_000, 0.096, 0.075, 0.05, 1.073,  3_200_000,    600_000,  2_200_000,   300_000,  4_650_000, 0.068,  0.028),
    _build_performance_row("2017-12-31", 408_000_000, 47_800_000, 0.100, 0.080, 0.06, 1.086,  3_800_000,    600_000,  2_600_000,   320_000,  4_780_000, 0.069,  0.031),
    # 2018 — SOFR rising (4 qtrs)
    _build_performance_row("2018-03-31", 409_500_000, 48_300_000, 0.103, 0.083, 0.07, 1.098,  4_300_000,    500_000,  2_900_000,   340_000,  5_180_000, 0.071,  0.032),
    _build_performance_row("2018-06-30", 411_000_000, 48_700_000, 0.107, 0.087, 0.08, 1.107,  4_700_000,    400_000,  3_100_000,   380_000,  5_380_000, 0.073,  0.034),
    _build_performance_row("2018-09-30", 412_000_000, 49_100_000, 0.109, 0.090, 0.09, 1.116,  5_100_000,    400_000,  3_400_000,   420_000,  5_520_000, 0.075,  0.034),
    _build_performance_row("2018-12-31", 410_000_000, 48_500_000, 0.108, 0.089, 0.10, 1.102,  4_500_000,   -600_000,  2_800_000,   380_000,  5_480_000, 0.075,  0.033),
    # 2019 — Spread compression, last year of reinvestment (4 qtrs)
    _build_performance_row("2019-03-31", 412_000_000, 49_000_000, 0.110, 0.091, 0.11, 1.114,  5_000_000,    500_000,  3_200_000,   400_000,  5_600_000, 0.074,  0.036),
    _build_performance_row("2019-06-30", 413_500_000, 49_600_000, 0.113, 0.094, 0.12, 1.127,  5_600_000,    600_000,  3_600_000,   420_000,  5_680_000, 0.073,  0.040),
    _build_performance_row("2019-09-30", 415_000_000, 50_200_000, 0.116, 0.097, 0.13, 1.141,  6_200_000,    600_000,  4_000_000,   450_000,  5_740_000, 0.072,  0.044),
    _build_performance_row("2019-12-31", 417_500_000, 51_000_000, 0.119, 0.100, 0.15, 1.159,  7_000_000,    800_000,  4_600_000,   500_000,  5_800_000, 0.072,  0.047),
    # 2020 — COVID shock; reinvestment ends Apr 2020 (4 qtrs)
    _build_performance_row("2020-03-31", 372_000_000, 38_500_000, 0.108, 0.089, 0.16, 0.875, -1_500_000, -8_500_000, -4_000_000,   300_000,  3_580_000, 0.057,  0.051),
    _build_performance_row("2020-06-30", 382_000_000, 40_500_000, 0.108, 0.090, 0.17, 0.920,  1_000_000,  2_500_000, -1_500_000,   400_000,  3_620_000, 0.057,  0.051),
    _build_performance_row("2020-09-30", 405_000_000, 46_000_000, 0.113, 0.094, 0.18, 1.045,  5_800_000,  4_800_000,  2_800_000,   450_000,  3_680_000, 0.058,  0.055),
    _build_performance_row("2020-12-31", 413_000_000, 49_200_000, 0.118, 0.099, 0.20, 1.118,  9_800_000,  4_000_000,  5_800_000,   500_000,  3_740_000, 0.059,  0.059),
    # 2021 — Low rates, fund amortising (4 qtrs)
    _build_performance_row("2021-03-31", 415_000_000, 50_000_000, 0.121, 0.102, 0.22, 1.136, 10_800_000,  1_000_000,  6_400_000,   550_000,  3_780_000, 0.060,  0.061),
    _build_performance_row("2021-06-30", 417_000_000, 50_600_000, 0.124, 0.104, 0.24, 1.150, 11_400_000,    600_000,  6_800_000,   580_000,  3_820_000, 0.061,  0.063),
    _build_performance_row("2021-09-30", 419_000_000, 51_200_000, 0.126, 0.106, 0.26, 1.164, 12_000_000,    600_000,  7_200_000,   620_000,  3_850_000, 0.061,  0.065),
    _build_performance_row("2021-12-31", 420_000_000, 51_800_000, 0.128, 0.108, 0.28, 1.177, 12_800_000,    800_000,  7_800_000,   660_000,  3_880_000, 0.062,  0.066),
    # 2022 — Aggressive Fed rate hikes (4 qtrs)
    _build_performance_row("2022-03-31", 418_000_000, 51_200_000, 0.129, 0.109, 0.34, 1.164, 13_200_000,    400_000,  8_000_000,   700_000,  5_200_000, 0.070,  0.059),
    _build_performance_row("2022-06-30", 408_000_000, 49_000_000, 0.129, 0.109, 0.43, 1.114, 11_000_000, -2_200_000,  5_800_000,   680_000,  7_200_000, 0.080,  0.049),
    _build_performance_row("2022-09-30", 395_000_000, 47_000_000, 0.130, 0.110, 0.51, 1.068,  9_800_000, -1_200_000,  4_600_000,   660_000,  8_400_000, 0.085,  0.045),
    _build_performance_row("2022-12-31", 385_000_000, 46_500_000, 0.131, 0.111, 0.57, 1.057, 10_500_000,    700_000,  5_200_000,   680_000,  8_200_000, 0.085,  0.046),
    # 2023 — High SOFR, significant amortisation (4 qtrs)
    _build_performance_row("2023-03-31", 368_000_000, 46_000_000, 0.133, 0.112, 0.68, 1.045, 12_000_000,  1_500_000,  6_000_000,   700_000,  7_500_000, 0.088,  0.045),
    _build_performance_row("2023-06-30", 348_000_000, 45_500_000, 0.134, 0.113, 0.78, 1.034, 13_500_000,  1_500_000,  6_500_000,   720_000,  7_200_000, 0.090,  0.044),
    _build_performance_row("2023-09-30", 325_000_000, 45_000_000, 0.135, 0.114, 0.88, 1.023, 15_000_000,  1_500_000,  7_000_000,   750_000,  6_800_000, 0.091,  0.044),
    _build_performance_row("2023-12-31", 305_000_000, 44_500_000, 0.136, 0.115, 0.98, 1.011, 16_500_000,  1_500_000,  7_500_000,   780_000,  6_500_000, 0.092,  0.044),
    # 2024 — Final years, rapid amortisation (4 qtrs)
    _build_performance_row("2024-03-31", 285_000_000, 43_800_000, 0.137, 0.116, 1.08, 0.995, 17_800_000,  1_300_000,  8_000_000,   800_000,  6_200_000, 0.090,  0.047),
    _build_performance_row("2024-06-30", 270_000_000, 43_200_000, 0.138, 0.116, 1.17, 0.982, 19_000_000,  1_200_000,  8_400_000,   820_000,  5_980_000, 0.089,  0.049),
    _build_performance_row("2024-09-30", 255_000_000, 42_500_000, 0.139, 0.117, 1.26, 0.966, 20_000_000,  1_000_000,  8_600_000,   840_000,  5_750_000, 0.088,  0.051),
    _build_performance_row("2024-12-31", 240_000_000, 41_800_000, 0.139, 0.117, 1.35, 0.950, 21_200_000,  1_200_000,  9_000_000,   860_000,  5_500_000, 0.088,  0.051),
    # 2025 — Near final (4 qtrs)
    _build_performance_row("2025-03-31", 225_000_000, 40_500_000, 0.140, 0.118, 1.44, 0.920, 22_000_000,    800_000,  9_200_000,   880_000,  5_200_000, 0.087,  0.053),
    _build_performance_row("2025-06-30", 212_000_000, 39_500_000, 0.141, 0.118, 1.53, 0.898, 23_000_000,  1_000_000,  9_500_000,   900_000,  4_980_000, 0.086,  0.055),
    _build_performance_row("2025-09-30", 198_000_000, 38_500_000, 0.141, 0.118, 1.62, 0.875, 24_000_000,  1_000_000,  9_800_000,   920_000,  4_760_000, 0.086,  0.055),
    _build_performance_row("2025-12-31", 184_000_000, 37_500_000, 0.142, 0.118, 1.72, 0.852, 25_000_000,  1_000_000, 10_000_000,   940_000,  4_540_000, 0.087,  0.055),
    # 2026 Q1 — Current reporting date
    _build_performance_row("2026-03-31", 174_500_000, 37_200_000, 0.142, 0.118, 1.78, 0.845, 25_800_000,    800_000, 10_200_000,   960_000,  4_320_000, 0.088,  0.054),
]

# DP-04 — Compliance Dashboard (amortising fund, better OC cushions)
F2_DP04_ROWS = [
    ("OC-A", "Class A/B OC Test",        "OC",          "A/B",  1.512, 1.245, 0.267, "PASS",
     "Divert principal & interest proceeds to repay Class A until cured",   FUND2_REPORTING_DATE),
    ("OC-C", "Class C OC Test",          "OC",          "C",    1.428, 1.170, 0.258, "PASS",
     "Divert proceeds to repay Class A then Class B principal",             FUND2_REPORTING_DATE),
    ("OC-D", "Class D OC Test",          "OC",          "D",    1.352, 1.105, 0.247, "PASS",
     "Divert proceeds to repay senior tranches in priority order",          FUND2_REPORTING_DATE),
    ("OC-E", "Class E OC Test",          "OC",          "E",    1.272, 1.055, 0.217, "PASS",
     "Divert proceeds to repay senior tranches in priority order",          FUND2_REPORTING_DATE),
    ("IC-A", "Class A/B IC Test",        "IC",          "A/B",  2.180, 1.200, 0.980, "PASS",
     "Divert proceeds to repay Class A principal",                          FUND2_REPORTING_DATE),
    ("IC-C", "Class C IC Test",          "IC",          "C",    1.920, 1.150, 0.770, "PASS",
     "Divert proceeds to repay senior tranches",                            FUND2_REPORTING_DATE),
    ("IC-D", "Class D IC Test",          "IC",          "D",    1.680, 1.105, 0.575, "PASS",
     "Divert proceeds to repay senior tranches",                            FUND2_REPORTING_DATE),
    ("IC-E", "Class E IC Test",          "IC",          "E",    1.450, 1.085, 0.365, "PASS",
     "Divert proceeds to repay senior tranches",                            FUND2_REPORTING_DATE),
    ("WARF", "Moody's WARF Covenant",    "Quality",     "Fund", 2840,  2900,   60,   "PASS",
     "Trading restrictions; CCC bucket reclassification",                   FUND2_REPORTING_DATE),
    ("DIV",  "Diversity Score",          "Quality",     "Fund", 72.0,  70.0,  2.0,   "PASS",
     "Trading restrictions; new purchase ineligibility",                    FUND2_REPORTING_DATE),
    ("CCC",  "CCC/Caa Bucket %",         "Concentration","Fund",0.078, 0.075,-0.003, "FAIL",
     "Excess CCC haircut to market value in OC test",                       FUND2_REPORTING_DATE),
    ("OBL",  "Largest Single Obligor %", "Concentration","Fund",0.080, 0.100, 0.020, "PASS",
     "Trading restrictions until cured",                                    FUND2_REPORTING_DATE),
    ("IND",  "Largest Single Industry %","Concentration","Fund",0.180, 0.200, 0.020, "PASS",
     "Trading restrictions until cured",                                    FUND2_REPORTING_DATE),
    ("PIK",  "DIP/PIK %",               "Concentration","Fund",0.057, 0.075, 0.018, "PASS",
     "Excess treated as defaulted in OC test",                              FUND2_REPORTING_DATE),
]

# DP-05 — Cashflow Statement (last 4 qtrs, larger principal repayments)
def _f2_waterfall(
    payment_date: str,
    collection_period: str,
    interest: float,
    principal: float,
    equity_dist: float,
    mgmt_fee: float,
    incentive_fee: float,
    trustee_fee: float,
) -> list[tuple]:
    rows = []
    base = (payment_date, collection_period, interest, principal, 0, 0)
    a_interest   =   810_000
    bcde_interest =  138_750 + 119_438 + 134_400 + 235_600
    sub_interest  =   385_000
    rows.append((*base, "1. Senior Expenses",           "Trustee + Admin + Tax",           trustee_fee + 60_000, 0, 0, 0, 0, trustee_fee))
    rows.append((*base, "2. Class A Interest",          "Class A Tranche",                 a_interest,           0, 0, 0, 0, 0))
    rows.append((*base, "3. Class A Principal Repay",   "Class A Tranche (amortisation)",  principal,            0, 0, 0, 0, 0))
    rows.append((*base, "4. Class A OC Test",           "Test passes — no diversion",      0,                    0, 0, 0, 0, 0))
    rows.append((*base, "5. Class B–E Interest & Tests","Class B/C/D/E Tranches",          bcde_interest,        0, 0, 0, 0, 0))
    rows.append((*base, "6. Senior Management Fee",     "DKIG Asset Management",           mgmt_fee * 0.375,     0, 0, mgmt_fee * 0.375, 0, 0))
    rows.append((*base, "7. Subordinated Notes Interest","Sub Notes Holders",              sub_interest,         0, 0, 0, 0, 0))
    rows.append((*base, "8. Incentive Fee / Sub Mgmt Fee","DKIG Asset Management",         incentive_fee + mgmt_fee * 0.625, 0, 0, mgmt_fee * 0.625, incentive_fee, 0))
    rows.append((*base, "9. Equity Distribution",       "Preference Shareholders (Equity)",equity_dist,         0, equity_dist, 0, 0, 0))
    return rows

F2_DP05_ROWS = []
F2_DP05_ROWS += _f2_waterfall("2025-07-21","2025-04-21 → 2025-07-20",
                               4_980_000, 18_000_000, equity_dist=1_250_000,
                               mgmt_fee=262_500, incentive_fee=0, trustee_fee=32_500)
F2_DP05_ROWS += _f2_waterfall("2025-10-21","2025-07-21 → 2025-10-20",
                               4_760_000, 19_500_000, equity_dist=1_100_000,
                               mgmt_fee=250_000, incentive_fee=0, trustee_fee=32_500)
F2_DP05_ROWS += _f2_waterfall("2026-01-21","2025-10-21 → 2026-01-20",
                               4_540_000, 21_000_000, equity_dist=1_050_000,
                               mgmt_fee=237_500, incentive_fee=0, trustee_fee=32_500)
F2_DP05_ROWS += _f2_waterfall("2026-04-21","2026-01-21 → 2026-04-20",
                               4_320_000, 22_500_000, equity_dist=  980_000,
                               mgmt_fee=225_000, incentive_fee=0, trustee_fee=32_500)

# DP-06 — Fee & Expense Ledger (smaller base on amortised portfolio)
F2_DP06_ROWS = [
    ("2026 Q1","Management Fee — Senior",       "0.15% p.a.",         196_875,   65_625,   65_625, 5_512_500, None, None,  None,    None,   None),
    ("2026 Q1","Management Fee — Subordinated", "0.25% p.a.",         328_125,  109_375,  109_375, 9_187_500, None, None,  None,    None,   None),
    ("2026 Q1","Incentive Fee",                 "20% above 12% IRR",        0,        0,        0,         0, 0.12, 0.20,  None,    None,   None),
    ("2026 Q1","Trustee Fee",                   "USD 130,000 p.a.",    32_500,   32_500,   32_500,   910_000, None, None,  None,    None,   None),
    ("2026 Q1","Admin / Accounting Fee",        "USD 80,000 p.a.",     20_000,   20_000,   20_000,   560_000, None, None,  None,    None,   None),
    ("2026 Q1","Legal Fee",                     "Variable",            18_000,   18_000,   12_000,   504_000, None, None,  None,    None,   None),
    ("2026 Q1","Rating Agency Fee",             "USD 180,000 p.a.",    45_000,   45_000,   45_000, 1_260_000, None, None,  None,    None,   None),
    ("2026 Q1","Tax Provision",                 "Effective",                0,        0,        0,         0, None, None, 145_000, 0.0185, None),
    ("2026 YTD","Total Expense Ratio (TER)",    "Aggregate",          640_500,  290_500,  284_500,17_934_000, None, None, 145_000, 0.0185, 0.0138),
]

# DP-07 — Key Metrics (last 8 weeks; shorter WAL, lower diversity — fewer positions)
F2_DP07_ROWS = [
    ("2026-02-02", 380, 2830, 3.2, 0.095, 0.60, -800_000, 1.00, 0.00, 0.052, 0.076, 0.73, 73.0, 78, 21, 0.080, 0.180, 0.318),
    ("2026-02-09", 381, 2835, 3.1, 0.095, 0.60, -750_000, 1.00, 0.00, 0.054, 0.077, 0.73, 73.0, 78, 21, 0.080, 0.180, 0.319),
    ("2026-02-16", 382, 2838, 3.1, 0.096, 0.60, -700_000, 1.00, 0.00, 0.055, 0.077, 0.73, 72.5, 78, 21, 0.080, 0.180, 0.320),
    ("2026-02-23", 383, 2840, 3.0, 0.096, 0.60, -660_000, 1.00, 0.00, 0.056, 0.078, 0.73, 72.5, 77, 21, 0.081, 0.180, 0.321),
    ("2026-03-02", 383, 2840, 3.0, 0.096, 0.60, -620_000, 1.00, 0.00, 0.057, 0.078, 0.73, 72.5, 77, 21, 0.081, 0.181, 0.321),
    ("2026-03-09", 384, 2842, 2.9, 0.097, 0.60, -580_000, 1.00, 0.00, 0.057, 0.078, 0.73, 72.0, 77, 21, 0.081, 0.181, 0.322),
    ("2026-03-16", 384, 2840, 2.9, 0.097, 0.60, -540_000, 1.00, 0.00, 0.057, 0.078, 0.73, 72.0, 77, 21, 0.081, 0.181, 0.322),
    ("2026-03-31", 385, 2840, 2.8, 0.097, 0.60, -500_000, 1.00, 0.00, 0.057, 0.078, 0.73, 72.0, 77, 21, 0.081, 0.181, 0.322),
]

# ---------------------------------------------------------------------------
# Cover sheet
# ---------------------------------------------------------------------------
def write_cover(wb) -> None:
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 70

    ws.cell(row=1, column=1, value="CLO Fund — Domain Data").font = Font(name="Bitter", size=22, bold=True, color=COLORS["primary"])
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 36

    ws.cell(row=2, column=1, value="Synthetic data for two funds covering all 8 data products defined in the DKIG Domain Ontology.").font = SUBTITLE_FONT
    ws.merge_cells("A2:B2")

    ws.cell(row=4, column=1, value="Funds").font = Font(name="Inter", size=10, bold=True, color=COLORS["primary"])
    ws.cell(row=4, column=2, value="DKIG Funding 2024-VII LLC  ·  DKIG Funding 2016-I LLC").font = BODY_FONT
    ws.cell(row=5, column=1, value="Manager").font = Font(name="Inter", size=10, bold=True, color=COLORS["primary"])
    ws.cell(row=5, column=2, value="DKIG Asset Management LLC").font = BODY_FONT
    ws.cell(row=6, column=1, value="Classification").font = Font(name="Inter", size=10, bold=True, color=COLORS["primary"])
    ws.cell(row=6, column=2, value="Confidential — Synthetic Sample Data").font = BODY_FONT
    ws.cell(row=7, column=1, value="Companion Ontology").font = Font(name="Inter", size=10, bold=True, color=COLORS["primary"])
    ws.cell(row=7, column=2, value="clo-fund-ontology.jsonld").font = BODY_FONT

    # Sheet index
    ws.cell(row=10, column=1, value="Sheet Index").font = Font(name="Bitter", size=14, bold=True, color=COLORS["primary"])
    sheets = [
        ("DP-01 Static Profile",      "Static",   "2024-VII — Legal structure, indenture, objectives"),
        ("DP-02 Portfolio Snapshot",  "Dynamic",  "2024-VII — Loan-level holdings, prices, ratings"),
        ("DP-03 Performance",         "Dynamic",  "2024-VII — NAV, IRR, DPI/RVPI/TVPI (12 months)"),
        ("DP-04 Compliance",          "Dynamic",  "2024-VII — OC/IC/Diversity/Concentration tests"),
        ("DP-05 Cashflows",           "Dynamic",  "2024-VII — Waterfall cashflows (last 4 qtrs)"),
        ("DP-06 Fees",                "Dynamic",  "2024-VII — Fee & expense accruals"),
        ("DP-07 Key Metrics",         "Dynamic",  "2024-VII — WAS, WARF, WAL, diversity"),
        ("DP-08 Liability Structure", "Mixed",    "2024-VII — Tranche stack, ratings, notional"),
        ("2016-DP-01 Static Profile", "Static",   "2016-I — Legal structure, indenture, objectives"),
        ("2016-DP-02 Portfolio Snapshot","Dynamic","2016-I — Current portfolio (~$175M par, 15 positions)"),
        ("2016-DP-03 Performance",    "Dynamic",  "2016-I — 40 quarters Q2-2016 → Q1-2026"),
        ("2016-DP-04 Compliance",     "Dynamic",  "2016-I — OC/IC tests (1 CCC breach flagged)"),
        ("2016-DP-05 Cashflows",      "Dynamic",  "2016-I — Waterfall + principal amortisation"),
        ("2016-DP-06 Fees",           "Dynamic",  "2016-I — Fee & expense accruals"),
        ("2016-DP-07 Key Metrics",    "Dynamic",  "2016-I — WAS, WARF, WAL (shorter, fewer positions)"),
        ("2016-DP-08 Liability Structure","Mixed", "2016-I — Tranche stack (Class A 70% amortised)"),
    ]
    headers = ["Sheet", "Type", "Description"]
    for c, label in enumerate(headers, start=1):
        cell = ws.cell(row=14, column=c, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
    for i, (s, t, d) in enumerate(sheets, start=12):
        is_band = (i - 11) % 2 == 0
        for c, value in enumerate([s, t, d], start=1):
            cell = ws.cell(row=i, column=c, value=value)
            cell.font = BAND_FONT
            cell.border = BORDER
            if is_band:
                cell.fill = BAND_FILL
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 62


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------
wb = Workbook()
write_cover(wb)
write_sheet(wb, "DP-01 Static Profile", "DP-01", "Fund Static Profile",
            "Immutable reference data — legal structure, indenture, objectives.",
            DP01_COLS, DP01_ROWS, number_formats=DP01_FORMATS)
write_sheet(wb, "DP-02 Portfolio Snapshot", "DP-02", "Fund Portfolio Snapshot",
            f"Loan-level positions held by the fund as of {REPORTING_DATE}.",
            DP02_COLS, DP02_ROWS, number_formats=DP02_FORMATS)
write_sheet(wb, "DP-03 Performance", "DP-03", "Fund Performance Metrics",
            "Trailing 12-month performance history.",
            DP03_COLS, DP03_ROWS, number_formats=DP03_FORMATS)
write_sheet(wb, "DP-04 Compliance", "DP-04", "Fund Compliance Dashboard",
            f"OC, IC and concentration covenant status as of {REPORTING_DATE}.",
            DP04_COLS, DP04_ROWS, pass_col=7, number_formats=DP04_FORMATS)
write_sheet(wb, "DP-05 Cashflows", "DP-05", "Fund Cashflow Statement",
            "Trailing 4 quarterly payment-date waterfalls.",
            DP05_COLS, DP05_ROWS, number_formats=DP05_FORMATS)
write_sheet(wb, "DP-06 Fees", "DP-06", "Fund Fee & Expense Ledger",
            "Q1 2026 fee, expense and tax accruals.",
            DP06_COLS, DP06_ROWS, number_formats=DP06_FORMATS)
write_sheet(wb, "DP-07 Key Metrics", "DP-07", "Fund Key Metrics Tracker",
            "Weekly portfolio-quality metrics over the last 8 weeks.",
            DP07_COLS, DP07_ROWS, number_formats=DP07_FORMATS)
write_sheet(wb, "DP-08 Liability Structure", "DP-08", "Fund Liability Structure",
            "Tranche stack — current notional, ratings, OC/IC cushion.",
            DP08_COLS, DP08_ROWS, number_formats=DP08_FORMATS)

# ---- Fund 2: DKIG Funding 2016-I LLC ----
write_sheet(wb, "2016-DP-01 Static Profile", "DP-01", "Fund Static Profile — DKIG 2016-I",
            "Immutable reference data — legal structure, indenture, objectives.",
            DP01_COLS, F2_DP01_ROWS, number_formats=DP01_FORMATS)
write_sheet(wb, "2016-DP-02 Portfolio Snapshot", "DP-02", "Fund Portfolio Snapshot — DKIG 2016-I",
            f"Loan-level positions held by the fund as of {FUND2_REPORTING_DATE} (~$175M par remaining).",
            DP02_COLS, F2_DP02_ROWS, number_formats=DP02_FORMATS)
write_sheet(wb, "2016-DP-03 Performance", "DP-03", "Fund Performance Metrics — DKIG 2016-I",
            "Quarterly performance history Q2-2016 → Q1-2026 (40 quarters since inception).",
            DP03_COLS, F2_DP03_ROWS, number_formats=DP03_FORMATS)
write_sheet(wb, "2016-DP-04 Compliance", "DP-04", "Fund Compliance Dashboard — DKIG 2016-I",
            f"OC, IC and concentration covenant status as of {FUND2_REPORTING_DATE}. Note: CCC bucket in breach.",
            DP04_COLS, F2_DP04_ROWS, pass_col=7, number_formats=DP04_FORMATS)
write_sheet(wb, "2016-DP-05 Cashflows", "DP-05", "Fund Cashflow Statement — DKIG 2016-I",
            "Last 4 quarterly payment-date waterfalls including principal amortisation.",
            DP05_COLS, F2_DP05_ROWS, number_formats=DP05_FORMATS)
write_sheet(wb, "2016-DP-06 Fees", "DP-06", "Fund Fee & Expense Ledger — DKIG 2016-I",
            "Q1 2026 fee, expense and tax accruals (smaller base — amortising portfolio).",
            DP06_COLS, F2_DP06_ROWS, number_formats=DP06_FORMATS)
write_sheet(wb, "2016-DP-07 Key Metrics", "DP-07", "Fund Key Metrics Tracker — DKIG 2016-I",
            "Weekly portfolio-quality metrics over the last 8 weeks.",
            DP07_COLS, F2_DP07_ROWS, number_formats=DP07_FORMATS)
write_sheet(wb, "2016-DP-08 Liability Structure", "DP-08", "Fund Liability Structure — DKIG 2016-I",
            "Tranche stack — Class A 70% amortised; current notional, ratings, OC/IC cushion.",
            DP08_COLS, F2_DP08_ROWS, number_formats=DP08_FORMATS)

wb.save(XLSX_PATH)
print(f"[ok] wrote {XLSX_PATH.name}")

# ---------------------------------------------------------------------------
# JSON-LD ontology (companion)
# ---------------------------------------------------------------------------
ontology = {
    "@context": {
        "clo":  "https://cifc.com/ontology/clo-fund#",
        "dp":   "https://cifc.com/ontology/clo-fund/data-product#",
        "owl":  "http://www.w3.org/2002/07/owl#",
        "rdfs": "http://www.w3.org/2000/01/rdf-schema#",
        "xsd":  "http://www.w3.org/2001/XMLSchema#",
        "label":      "rdfs:label",
        "definition": "rdfs:comment",
        "subClassOf": {"@id": "rdfs:subClassOf", "@type": "@id"},
        "domain":     {"@id": "rdfs:domain", "@type": "@id"},
        "range":      {"@id": "rdfs:range", "@type": "@id"},
    },
    "@graph": [
        {"@id": "clo:CLOFund-Ontology",
         "@type": "owl:Ontology",
         "label": "DKIG CLO Fund Domain Ontology",
         "definition": "Formal vocabulary enabling AI agents to interpret and query CLO fund data products. Companion to CLO_Fund_Domain_Data.xlsx.",
         "owl:versionInfo": "1.0.0",
         "dataReportingDate": REPORTING_DATE},

        # ---------- Core classes ----------
        {"@id": "clo:CLOFund", "@type": "owl:Class", "label": "CLO Fund",
         "definition": "A Collateralised Loan Obligation fund — a special purpose vehicle that holds a portfolio of loans and issues tranched debt and equity to investors.",
         "subClassOf": "owl:Thing"},
        {"@id": "clo:FundIndenture", "@type": "owl:Class", "label": "Fund Indenture",
         "definition": "Legal document governing the fund's operation — defines tests, covenants, fees, waterfall, and eligible asset criteria.",
         "subClassOf": "owl:Thing"},
        {"@id": "clo:FundObjective", "@type": "owl:Class", "label": "Fund Objective",
         "definition": "Investment mandate and return objective of the fund.",
         "subClassOf": "owl:Thing"},
        {"@id": "clo:Tranche", "@type": "owl:Class", "label": "Tranche",
         "definition": "A class of debt issued by the CLO fund with defined seniority, coupon, and rating.",
         "subClassOf": "owl:Thing"},
        {"@id": "clo:SeniorTranche", "@type": "owl:Class", "label": "Senior Tranche",
         "definition": "The most senior-ranking debt class (Class A), first to receive interest and principal payments.",
         "subClassOf": "clo:Tranche"},
        {"@id": "clo:MezzanineTranche", "@type": "owl:Class", "label": "Mezzanine Tranche",
         "definition": "Debt classes ranking below senior (Class B/C/D/E) — higher yield, higher risk.",
         "subClassOf": "clo:Tranche"},
        {"@id": "clo:SubordinatedNotes", "@type": "owl:Class", "label": "Subordinated Notes",
         "definition": "Most junior debt tranche, immediately above equity in the waterfall.",
         "subClassOf": "clo:Tranche"},
        {"@id": "clo:EquityPiece", "@type": "owl:Class", "label": "Equity / Preference Shares",
         "definition": "Residual interest — receives cashflows after all debt tranches are paid. Bears first loss.",
         "subClassOf": "owl:Thing"},
        {"@id": "clo:LoanAsset", "@type": "owl:Class", "label": "Loan Asset",
         "definition": "Debt instrument held as collateral in the fund's asset pool.",
         "subClassOf": "owl:Thing"},
        {"@id": "clo:SyndicatedLoan", "@type": "owl:Class", "label": "Syndicated Loan",
         "definition": "Broadly syndicated senior secured leveraged loan — primary asset type in a CLO.",
         "subClassOf": "clo:LoanAsset"},
        {"@id": "clo:PrivateCreditLoan", "@type": "owl:Class", "label": "Private Credit Loan",
         "definition": "Directly originated, privately negotiated loan — middle-market or unitranche facilities.",
         "subClassOf": "clo:LoanAsset"},
        {"@id": "clo:Obligor", "@type": "owl:Class", "label": "Obligor",
         "definition": "Corporate borrower that issued the loan asset.",
         "subClassOf": "owl:Thing"},
        {"@id": "clo:Waterfall", "@type": "owl:Class", "label": "Payment Waterfall",
         "definition": "Contractually defined priority sequence in which cashflows are distributed.",
         "subClassOf": "owl:Thing"},
        {"@id": "clo:ComplianceTest", "@type": "owl:Class", "label": "Compliance Test",
         "definition": "Coverage or quality test defined in the indenture that must be satisfied for equity distributions to continue.",
         "subClassOf": "owl:Thing"},
        {"@id": "clo:OCTest", "@type": "owl:Class", "label": "OC Test",
         "definition": "Overcollateralisation test — checks if par value of collateral sufficiently exceeds tranche notional.",
         "subClassOf": "clo:ComplianceTest"},
        {"@id": "clo:ICTest", "@type": "owl:Class", "label": "IC Test",
         "definition": "Interest Coverage test — checks if interest income covers interest payable on a tranche class.",
         "subClassOf": "clo:ComplianceTest"},
        {"@id": "clo:DiversityTest", "@type": "owl:Class", "label": "Diversity Test",
         "definition": "Tests whether portfolio is sufficiently diversified per Moody's diversity score methodology.",
         "subClassOf": "clo:ComplianceTest"},
        {"@id": "clo:CashflowEvent", "@type": "owl:Class", "label": "Cashflow Event",
         "definition": "A discrete cash movement — income from assets or disbursements to liabilities.",
         "subClassOf": "owl:Thing"},
        {"@id": "clo:InterestPayment", "@type": "owl:Class", "label": "Interest Payment",
         "definition": "Cash received from an obligor as scheduled interest on a loan.",
         "subClassOf": "clo:CashflowEvent"},
        {"@id": "clo:PrincipalPayment", "@type": "owl:Class", "label": "Principal Payment",
         "definition": "Cash received as scheduled or unscheduled repayment of loan principal.",
         "subClassOf": "clo:CashflowEvent"},
        {"@id": "clo:TrancheDistribution", "@type": "owl:Class", "label": "Tranche Distribution",
         "definition": "Cash disbursed to a tranche class as interest or principal per the waterfall.",
         "subClassOf": "clo:CashflowEvent"},
        {"@id": "clo:EquityDistribution", "@type": "owl:Class", "label": "Equity Distribution",
         "definition": "Residual cash disbursed to equity holders after all senior obligations are met.",
         "subClassOf": "clo:CashflowEvent"},
        {"@id": "clo:FeeExpense", "@type": "owl:Class", "label": "Fee / Expense",
         "definition": "Cost incurred by the fund — management, incentive, trustee, taxes.",
         "subClassOf": "owl:Thing"},
        {"@id": "clo:ManagementFee", "@type": "owl:Class", "label": "Management Fee",
         "definition": "Fee paid to the CLO manager — typically a fixed percentage of fund assets.",
         "subClassOf": "clo:FeeExpense"},
        {"@id": "clo:IncentiveFee", "@type": "owl:Class", "label": "Incentive Fee",
         "definition": "Performance-based fee paid above a hurdle rate. Also called subordinated management fee.",
         "subClassOf": "clo:FeeExpense"},
        {"@id": "clo:PerformanceSnapshot", "@type": "owl:Class", "label": "Performance Snapshot",
         "definition": "Time-stamped record of fund-level performance (NAV, IRR, P&L, returns).",
         "subClassOf": "owl:Thing"},
        {"@id": "clo:KeyMetricSnapshot", "@type": "owl:Class", "label": "Key Metric Snapshot",
         "definition": "Time-stamped record of portfolio quality metrics (WARF, WAS, diversity, WAL).",
         "subClassOf": "owl:Thing"},
        {"@id": "clo:RatingAssessment", "@type": "owl:Class", "label": "Rating Assessment",
         "definition": "Credit rating assigned to a loan asset, tranche, or obligor by a rating agency.",
         "subClassOf": "owl:Thing"},
        {"@id": "clo:RatingAgency", "@type": "owl:Class", "label": "Rating Agency",
         "definition": "External credit rating organisation. Instances: Moody's, S&P, Fitch.",
         "subClassOf": "owl:Thing"},

        # ---------- Object properties ----------
        {"@id": "clo:hasAsset",         "@type": "owl:ObjectProperty", "label": "has asset",         "domain": "clo:CLOFund",   "range": "clo:LoanAsset",         "cardinality": "1..*", "definition": "The fund holds one or more loan assets in its collateral pool."},
        {"@id": "clo:hasTranche",       "@type": "owl:ObjectProperty", "label": "has tranche",       "domain": "clo:CLOFund",   "range": "clo:Tranche",           "cardinality": "2..*", "definition": "The fund issues two or more debt tranches to investors."},
        {"@id": "clo:hasEquity",        "@type": "owl:ObjectProperty", "label": "has equity",        "domain": "clo:CLOFund",   "range": "clo:EquityPiece",       "cardinality": "1",    "definition": "The fund has exactly one equity piece (preference shares)."},
        {"@id": "clo:governedBy",       "@type": "owl:ObjectProperty", "label": "governed by",       "domain": "clo:CLOFund",   "range": "clo:FundIndenture",     "cardinality": "1",    "definition": "The fund is legally governed by exactly one indenture."},
        {"@id": "clo:hasObjective",     "@type": "owl:ObjectProperty", "label": "has objective",     "domain": "clo:CLOFund",   "range": "clo:FundObjective",     "cardinality": "1",    "definition": "The fund has a stated investment objective."},
        {"@id": "clo:hasComplianceTest","@type": "owl:ObjectProperty", "label": "has compliance test","domain": "clo:CLOFund",  "range": "clo:ComplianceTest",    "cardinality": "1..*", "definition": "The fund is subject to one or more compliance tests."},
        {"@id": "clo:hasFee",           "@type": "owl:ObjectProperty", "label": "has fee",           "domain": "clo:CLOFund",   "range": "clo:FeeExpense",        "cardinality": "1..*", "definition": "The fund incurs one or more types of fees/expenses."},
        {"@id": "clo:hasPerformanceSnapshot","@type": "owl:ObjectProperty","label":"has performance snapshot","domain":"clo:CLOFund","range":"clo:PerformanceSnapshot","cardinality":"1..*","definition":"The fund has one performance snapshot per reporting period."},
        {"@id": "clo:hasKeyMetrics",    "@type": "owl:ObjectProperty", "label": "has key metrics",   "domain": "clo:CLOFund",   "range": "clo:KeyMetricSnapshot", "cardinality": "1..*", "definition": "The fund has one key metrics snapshot per reporting period."},
        {"@id": "clo:hasObligor",       "@type": "owl:ObjectProperty", "label": "has obligor",       "domain": "clo:LoanAsset", "range": "clo:Obligor",           "cardinality": "1",    "definition": "Each loan asset has exactly one obligor (borrower)."},
        {"@id": "clo:hasCashflow",      "@type": "owl:ObjectProperty", "label": "has cashflow",      "domain": "clo:LoanAsset", "range": "clo:CashflowEvent",     "cardinality": "0..*", "definition": "A loan asset generates zero or more cashflow events."},
        {"@id": "clo:hasRating",        "@type": "owl:ObjectProperty", "label": "has rating",        "domain": "clo:LoanAsset", "range": "clo:RatingAssessment",  "cardinality": "0..*", "definition": "A loan asset has ratings from one or more agencies."},
        {"@id": "clo:ratedBy",          "@type": "owl:ObjectProperty", "label": "rated by",          "domain": "clo:RatingAssessment","range":"clo:RatingAgency", "cardinality": "1",    "definition": "A rating assessment is issued by exactly one rating agency."},
        {"@id": "clo:seniorTo",         "@type": "owl:ObjectProperty", "label": "senior to",         "domain": "clo:Tranche",   "range": "clo:Tranche",           "cardinality": "0..*", "definition": "A tranche ranks senior to one or more other tranches in the waterfall."},
        {"@id": "clo:distributedVia",   "@type": "owl:ObjectProperty", "label": "distributed via",   "domain": "clo:CashflowEvent","range":"clo:Waterfall",       "cardinality": "0..1", "definition": "A liability cashflow is distributed via the waterfall mechanism."},
        {"@id": "clo:coveredByTest",    "@type": "owl:ObjectProperty", "label": "covered by test",   "domain": "clo:Tranche",   "range": "clo:ComplianceTest",    "cardinality": "1..*", "definition": "A tranche class is protected by one or more compliance tests."},
        {"@id": "clo:appliesToFund",    "@type": "owl:ObjectProperty", "label": "applies to fund",   "domain": "clo:ComplianceTest","range":"clo:CLOFund",        "cardinality": "1",    "definition": "A compliance test applies to exactly one fund."},
        {"@id": "clo:hasObligorRating", "@type": "owl:ObjectProperty", "label": "has obligor rating","domain": "clo:Obligor",   "range": "clo:RatingAssessment",  "cardinality": "0..*", "definition": "An obligor has ratings from credit agencies."},
        {"@id": "clo:distributesTo",    "@type": "owl:ObjectProperty", "label": "distributes to",    "domain": "clo:Waterfall", "range": "clo:Tranche",           "cardinality": "2..*", "definition": "The waterfall distributes proceeds to each tranche in priority order."},

        # ---------- Data properties (subset most queried) ----------
        {"@id": "clo:fundName",              "@type": "owl:DatatypeProperty", "label": "fund name",            "domain": "clo:CLOFund",            "range": "xsd:string",  "definition": "Legal name of the CLO fund."},
        {"@id": "clo:vintageYear",           "@type": "owl:DatatypeProperty", "label": "vintage year",         "domain": "clo:CLOFund",            "range": "xsd:integer", "definition": "Year the fund closed."},
        {"@id": "clo:targetPar",             "@type": "owl:DatatypeProperty", "label": "target par",           "domain": "clo:CLOFund",            "range": "xsd:decimal", "definition": "Target collateral par amount at close (USD)."},
        {"@id": "clo:reinvestmentPeriodEnd", "@type": "owl:DatatypeProperty", "label": "reinvestment period end","domain": "clo:CLOFund",          "range": "xsd:date",    "definition": "Date after which principal proceeds cannot be reinvested."},
        {"@id": "clo:nonCallPeriodEnd",      "@type": "owl:DatatypeProperty", "label": "non-call period end",  "domain": "clo:CLOFund",            "range": "xsd:date",    "definition": "Date before which the fund cannot be optionally redeemed."},
        {"@id": "clo:legalMaturity",         "@type": "owl:DatatypeProperty", "label": "legal maturity",       "domain": "clo:CLOFund",            "range": "xsd:date",    "definition": "Legal final maturity date of the fund."},
        {"@id": "clo:baseCurrency",          "@type": "owl:DatatypeProperty", "label": "base currency",        "domain": "clo:CLOFund",            "range": "xsd:string",  "definition": "Base currency of the fund (typically USD)."},
        {"@id": "clo:managerName",           "@type": "owl:DatatypeProperty", "label": "manager name",         "domain": "clo:CLOFund",            "range": "xsd:string",  "definition": "Name of the CLO collateral manager."},
        {"@id": "clo:parValue",              "@type": "owl:DatatypeProperty", "label": "par value",            "domain": "clo:LoanAsset",          "range": "xsd:decimal", "definition": "Par (face) value of the position held (USD)."},
        {"@id": "clo:marketValue",           "@type": "owl:DatatypeProperty", "label": "market value",         "domain": "clo:LoanAsset",          "range": "xsd:decimal", "definition": "Current market value of the position (USD)."},
        {"@id": "clo:price",                 "@type": "owl:DatatypeProperty", "label": "price",                "domain": "clo:LoanAsset",          "range": "xsd:decimal", "definition": "Current market price (% of par)."},
        {"@id": "clo:spread",                "@type": "owl:DatatypeProperty", "label": "spread",               "domain": "clo:LoanAsset",          "range": "xsd:decimal", "definition": "Spread over SOFR in basis points."},
        {"@id": "clo:maturityDate",          "@type": "owl:DatatypeProperty", "label": "maturity date",        "domain": "clo:LoanAsset",          "range": "xsd:date",    "definition": "Scheduled maturity date of the loan."},
        {"@id": "clo:loanType",              "@type": "owl:DatatypeProperty", "label": "loan type",            "domain": "clo:LoanAsset",          "range": "xsd:string",  "definition": "First lien, second lien, unitranche, revolver, etc."},
        {"@id": "clo:isPIK",                 "@type": "owl:DatatypeProperty", "label": "is PIK",               "domain": "clo:LoanAsset",          "range": "xsd:boolean", "definition": "Whether the loan pays interest in kind."},
        {"@id": "clo:isCovenantLite",        "@type": "owl:DatatypeProperty", "label": "is covenant-lite",     "domain": "clo:LoanAsset",          "range": "xsd:boolean", "definition": "Whether the loan lacks maintenance financial covenants."},
        {"@id": "clo:industryCode",          "@type": "owl:DatatypeProperty", "label": "industry code",        "domain": "clo:LoanAsset",          "range": "xsd:string",  "definition": "Industry classification per Moody's or S&P taxonomy."},
        {"@id": "clo:trancheClass",          "@type": "owl:DatatypeProperty", "label": "tranche class",        "domain": "clo:Tranche",            "range": "xsd:string",  "definition": "Class designation (A, B, C, D, E, Sub Notes)."},
        {"@id": "clo:initialNotional",       "@type": "owl:DatatypeProperty", "label": "initial notional",     "domain": "clo:Tranche",            "range": "xsd:decimal", "definition": "Original notional amount at issuance (USD)."},
        {"@id": "clo:currentNotional",       "@type": "owl:DatatypeProperty", "label": "current notional",     "domain": "clo:Tranche",            "range": "xsd:decimal", "definition": "Current outstanding notional after any amortisation."},
        {"@id": "clo:couponRate",            "@type": "owl:DatatypeProperty", "label": "coupon rate",          "domain": "clo:Tranche",            "range": "xsd:decimal", "definition": "Coupon rate in basis points (SOFR+)."},
        {"@id": "clo:subordinationLevel",    "@type": "owl:DatatypeProperty", "label": "subordination level",  "domain": "clo:Tranche",            "range": "xsd:decimal", "definition": "Percentage of fund assets subordinated to this tranche."},
        {"@id": "clo:waterfallPriority",     "@type": "owl:DatatypeProperty", "label": "waterfall priority",   "domain": "clo:Tranche",            "range": "xsd:integer", "definition": "Numeric rank in payment waterfall (1 = most senior)."},
        {"@id": "clo:testName",              "@type": "owl:DatatypeProperty", "label": "test name",            "domain": "clo:ComplianceTest",     "range": "xsd:string",  "definition": "Name of the test (e.g. 'Class A/B OC Test')."},
        {"@id": "clo:currentValue",          "@type": "owl:DatatypeProperty", "label": "current value",        "domain": "clo:ComplianceTest",     "range": "xsd:decimal", "definition": "Current computed ratio or score."},
        {"@id": "clo:threshold",             "@type": "owl:DatatypeProperty", "label": "threshold",            "domain": "clo:ComplianceTest",     "range": "xsd:decimal", "definition": "Minimum passing value per indenture."},
        {"@id": "clo:cushion",               "@type": "owl:DatatypeProperty", "label": "cushion",              "domain": "clo:ComplianceTest",     "range": "xsd:decimal", "definition": "Difference between current value and threshold (positive = passing)."},
        {"@id": "clo:passFail",              "@type": "owl:DatatypeProperty", "label": "pass/fail",            "domain": "clo:ComplianceTest",     "range": "xsd:boolean", "definition": "True if the test is currently passing."},
        {"@id": "clo:breachConsequence",     "@type": "owl:DatatypeProperty", "label": "breach consequence",   "domain": "clo:ComplianceTest",     "range": "xsd:string",  "definition": "What happens if the test fails (e.g. OC diversion, reinvestment suspension)."},
        {"@id": "clo:reportingDate",         "@type": "owl:DatatypeProperty", "label": "reporting date",       "domain": "clo:PerformanceSnapshot","range": "xsd:date",    "definition": "Date of the performance measurement."},
        {"@id": "clo:nav",                   "@type": "owl:DatatypeProperty", "label": "NAV",                  "domain": "clo:PerformanceSnapshot","range": "xsd:decimal", "definition": "Net Asset Value of the fund (USD)."},
        {"@id": "clo:equityNAV",             "@type": "owl:DatatypeProperty", "label": "equity NAV",           "domain": "clo:PerformanceSnapshot","range": "xsd:decimal", "definition": "NAV attributable to equity holders (USD)."},
        {"@id": "clo:grossIRR",              "@type": "owl:DatatypeProperty", "label": "gross IRR",            "domain": "clo:PerformanceSnapshot","range": "xsd:decimal", "definition": "Gross IRR inception to date (%)."},
        {"@id": "clo:netIRR",                "@type": "owl:DatatypeProperty", "label": "net IRR",              "domain": "clo:PerformanceSnapshot","range": "xsd:decimal", "definition": "Net IRR after fees inception to date (%)."},
        {"@id": "clo:dpi",                   "@type": "owl:DatatypeProperty", "label": "DPI",                  "domain": "clo:PerformanceSnapshot","range": "xsd:decimal", "definition": "Distributions to Paid-In ratio."},
        {"@id": "clo:rvpi",                  "@type": "owl:DatatypeProperty", "label": "RVPI",                 "domain": "clo:PerformanceSnapshot","range": "xsd:decimal", "definition": "Residual Value to Paid-In ratio."},
        {"@id": "clo:tvpi",                  "@type": "owl:DatatypeProperty", "label": "TVPI",                 "domain": "clo:PerformanceSnapshot","range": "xsd:decimal", "definition": "Total Value to Paid-In ratio (DPI + RVPI)."},
        {"@id": "clo:warf",                  "@type": "owl:DatatypeProperty", "label": "WARF",                 "domain": "clo:KeyMetricSnapshot",  "range": "xsd:decimal", "definition": "Weighted Average Rating Factor (Moody's)."},
        {"@id": "clo:weightedAverageSpread", "@type": "owl:DatatypeProperty", "label": "weighted avg spread",  "domain": "clo:KeyMetricSnapshot",  "range": "xsd:decimal", "definition": "Weighted average asset spread in bps over SOFR."},
        {"@id": "clo:weightedAverageLife",   "@type": "owl:DatatypeProperty", "label": "WAL",                  "domain": "clo:KeyMetricSnapshot",  "range": "xsd:decimal", "definition": "Weighted average life in years."},
        {"@id": "clo:diversityScore",        "@type": "owl:DatatypeProperty", "label": "diversity score",     "domain": "clo:KeyMetricSnapshot",  "range": "xsd:decimal", "definition": "Moody's diversity score for the portfolio."},

        # ---------- Business axioms ----------
        {"@id": "clo:WaterfallPriorityAxiom", "@type": "owl:NamedIndividual",
         "label": "Payment Waterfall Priority Order",
         "definition": "The contractual order in which collected proceeds are distributed each payment date.",
         "priorityOrder": [
             "1. Senior Expenses (Trustee Fees, Rating Agency Fees, Taxes, Administration)",
             "2. Class A Interest",
             "3. Class A OC Test — if failing, divert proceeds to repay Class A principal until test passes",
             "4. Class A IC Test — if failing, divert proceeds to repay Class A principal",
             "5. Class B Interest → Class B OC Test → Class B IC Test → repeat for C, D, E",
             "6. Senior Management Fee",
             "7. Subordinated Notes Interest",
             "8. Incentive Fee / Subordinated Management Fee",
             "9. Equity Distribution — residual to preference shareholders",
         ],
         "rule": "If any OC or IC test fails, cash is redirected to repay the most senior tranche principal. Equity receives nothing until all tests pass."},

        {"@id": "clo:OCTestFormulaAxiom", "@type": "owl:NamedIndividual",
         "label": "OC Test Formula & Adjustments",
         "definition": "How the overcollateralisation ratio is computed and adjusted.",
         "formula": "OC Ratio = Par Value of Eligible Collateral ÷ Notional of Covered Tranches",
         "cccAdjustment": "CCC/Caa-rated assets above the bucket threshold are haircut to their market value (not par) for OC calculation.",
         "defaultedAssetAdjustment": "Defaulted assets are valued at the lower of market value and assumed recovery rate.",
         "passCondition": "OC Ratio ≥ OC Threshold (typically 120–150% for senior classes per indenture).",
         "breachConsequence": "All principal and interest proceeds are diverted to repay the most senior outstanding tranche until the test passes. Equity distributions cease."},

        {"@id": "clo:ReinvestmentPeriodAxiom", "@type": "owl:NamedIndividual",
         "label": "Reinvestment Period Behaviour",
         "definition": "Rules governing how principal proceeds are handled before vs after the reinvestment period end date.",
         "duringReinvestment": "If clo:reinvestmentPeriodEnd > today, principal proceeds may be reinvested in new eligible collateral. Manager has discretion subject to eligibility tests.",
         "afterReinvestment": "Principal proceeds must be applied to repay tranches in waterfall order (senior first). Fund begins amortising. Manager can still trade for credit-quality maintenance but cannot grow the portfolio."},

        {"@id": "clo:RatingDerivedRulesAxiom", "@type": "owl:NamedIndividual",
         "label": "Rating-Derived Asset Classifications",
         "definition": "Classification rules an agent applies based on rating data.",
         "cccDefinition": "An asset rated Caa1 / CCC+ or below is classified as a CCC Asset.",
         "cccBucketRule": "CCC Assets exceeding the indenture threshold (typically 7.5% of portfolio par) are haircut to market value in OC test calculations.",
         "defaultedDefinition": "An asset is Defaulted if the obligor has filed for bankruptcy, missed a payment, or been classified as defaulted by the controlling rating agency.",
         "warfDefinition": "WARF = weighted average of Moody's rating factors by par value. Each rating maps to a numeric factor per Moody's table (e.g. Baa3 = 610, B1 = 940, Caa1 = 3490)."},

        # ---------- Data product registry ----------
        {"@id": "dp:DP-01", "@type": "dp:DataProduct",
         "label": "Fund Static Profile", "temporalNature": "Static",
         "sheet": "DP-01 Static Profile", "workbookFile": "CLO_Fund_Domain_Data.xlsx",
         "primaryClasses": ["clo:CLOFund", "clo:FundIndenture", "clo:FundObjective"],
         "updateFrequency": "One-time at fund close; amended only by indenture supplement",
         "qualitySLA": "100% completeness at fund close; zero-tolerance for indenture term errors",
         "sourceSystems": ["Intex (structure)", "Kanerai (indenture)", "Fund Administrator"]},
        {"@id": "dp:DP-02", "@type": "dp:DataProduct",
         "label": "Fund Portfolio Snapshot", "temporalNature": "Dynamic",
         "sheet": "DP-02 Portfolio Snapshot", "workbookFile": "CLO_Fund_Domain_Data.xlsx",
         "primaryClasses": ["clo:LoanAsset", "clo:Obligor", "clo:RatingAssessment"],
         "updateFrequency": "Daily (prices); Weekly (ratings, metrics); Monthly (full report)",
         "qualitySLA": "Prices T+0 EOD; Ratings within 24h of agency action; 99.9% completeness",
         "sourceSystems": ["Bloomberg", "LSTA", "Clear Par", "Moody's/S&P/Fitch", "Intex"]},
        {"@id": "dp:DP-03", "@type": "dp:DataProduct",
         "label": "Fund Performance Metrics", "temporalNature": "Dynamic",
         "sheet": "DP-03 Performance", "workbookFile": "CLO_Fund_Domain_Data.xlsx",
         "primaryClasses": ["clo:PerformanceSnapshot"],
         "updateFrequency": "Monthly (NAV, P&L); Quarterly (IRR, DPI, RVPI); Daily (estimated NAV)",
         "qualitySLA": "Monthly NAV within 10 business days of month end; Quarterly report within 30 days",
         "sourceSystems": ["Fund Administrator", "Bloomberg (benchmark)", "Internal OMS/PMS"]},
        {"@id": "dp:DP-04", "@type": "dp:DataProduct",
         "label": "Fund Compliance Dashboard", "temporalNature": "Dynamic",
         "sheet": "DP-04 Compliance", "workbookFile": "CLO_Fund_Domain_Data.xlsx",
         "primaryClasses": ["clo:ComplianceTest", "clo:OCTest", "clo:ICTest", "clo:DiversityTest"],
         "updateFrequency": "Daily (monitoring); Payment date (formal indenture determination)",
         "qualitySLA": "Test status updated T+1 after rating action or portfolio change",
         "sourceSystems": ["Intex (calculations)", "Kanerai (covenant monitoring)", "Internal risk systems"]},
        {"@id": "dp:DP-05", "@type": "dp:DataProduct",
         "label": "Fund Cashflow Statement", "temporalNature": "Dynamic",
         "sheet": "DP-05 Cashflows", "workbookFile": "CLO_Fund_Domain_Data.xlsx",
         "primaryClasses": ["clo:CashflowEvent", "clo:Waterfall", "clo:TrancheDistribution", "clo:EquityDistribution"],
         "updateFrequency": "Per payment period (typically quarterly); intra-period accruals daily",
         "qualitySLA": "Official waterfall report published within 2 business days of each payment date",
         "sourceSystems": ["Trustee", "Clear Par", "Bloomberg (accruals)"]},
        {"@id": "dp:DP-06", "@type": "dp:DataProduct",
         "label": "Fund Fee & Expense Ledger", "temporalNature": "Dynamic",
         "sheet": "DP-06 Fees", "workbookFile": "CLO_Fund_Domain_Data.xlsx",
         "primaryClasses": ["clo:FeeExpense", "clo:ManagementFee", "clo:IncentiveFee"],
         "updateFrequency": "Monthly (accruals); Per payment period (actual payments)",
         "qualitySLA": "Monthly accruals reconciled within 5 business days of month end",
         "sourceSystems": ["Fund Administrator", "Internal fee calculation engine"]},
        {"@id": "dp:DP-07", "@type": "dp:DataProduct",
         "label": "Fund Key Metrics Tracker", "temporalNature": "Dynamic",
         "sheet": "DP-07 Key Metrics", "workbookFile": "CLO_Fund_Domain_Data.xlsx",
         "primaryClasses": ["clo:KeyMetricSnapshot"],
         "updateFrequency": "Daily (price-sensitive); Weekly (full metric set)",
         "qualitySLA": "Daily metrics by 8am next business day; Weekly report Monday morning",
         "sourceSystems": ["Moody's (WARF)", "Bloomberg (prices/spreads)", "Aggregation from DP-02"]},
        {"@id": "dp:DP-08", "@type": "dp:DataProduct",
         "label": "Fund Liability Structure", "temporalNature": "Mixed",
         "sheet": "DP-08 Liability Structure", "workbookFile": "CLO_Fund_Domain_Data.xlsx",
         "primaryClasses": ["clo:Tranche", "clo:SeniorTranche", "clo:MezzanineTranche", "clo:SubordinatedNotes", "clo:EquityPiece"],
         "updateFrequency": "Static (structure); Ratings within 24h of agency action; Notional each payment date",
         "qualitySLA": "Structural data static at close; Ratings 24h SLA; Notional updated each payment date",
         "sourceSystems": ["Intex", "Kanerai", "Moody's/S&P/Fitch"]},

        # ---------- Class → DataProduct mapping for agent retrieval ----------
        {"@id": "clo:retrievalMap", "@type": "owl:NamedIndividual",
         "label": "Class to Data Product Retrieval Map",
         "definition": "For each ontology class, the data product(s) an agent should query to retrieve instances and current values.",
         "mappings": [
             {"class": "clo:CLOFund",             "dataProducts": ["dp:DP-01"]},
             {"class": "clo:FundIndenture",       "dataProducts": ["dp:DP-01"]},
             {"class": "clo:FundObjective",       "dataProducts": ["dp:DP-01"]},
             {"class": "clo:LoanAsset",           "dataProducts": ["dp:DP-02"]},
             {"class": "clo:Obligor",             "dataProducts": ["dp:DP-02"]},
             {"class": "clo:RatingAssessment",    "dataProducts": ["dp:DP-02", "dp:DP-08"]},
             {"class": "clo:Tranche",             "dataProducts": ["dp:DP-08"]},
             {"class": "clo:EquityPiece",         "dataProducts": ["dp:DP-08"]},
             {"class": "clo:ComplianceTest",      "dataProducts": ["dp:DP-04"]},
             {"class": "clo:OCTest",              "dataProducts": ["dp:DP-04"]},
             {"class": "clo:ICTest",              "dataProducts": ["dp:DP-04"]},
             {"class": "clo:DiversityTest",       "dataProducts": ["dp:DP-04"]},
             {"class": "clo:CashflowEvent",       "dataProducts": ["dp:DP-05"]},
             {"class": "clo:Waterfall",           "dataProducts": ["dp:DP-05"]},
             {"class": "clo:TrancheDistribution", "dataProducts": ["dp:DP-05"]},
             {"class": "clo:EquityDistribution",  "dataProducts": ["dp:DP-05"]},
             {"class": "clo:FeeExpense",          "dataProducts": ["dp:DP-06"]},
             {"class": "clo:ManagementFee",       "dataProducts": ["dp:DP-06"]},
             {"class": "clo:IncentiveFee",        "dataProducts": ["dp:DP-06"]},
             {"class": "clo:PerformanceSnapshot", "dataProducts": ["dp:DP-03"]},
             {"class": "clo:KeyMetricSnapshot",   "dataProducts": ["dp:DP-07"]},
         ]},

        # ---------- Sample agent query patterns ----------
        {"@id": "clo:queryPatterns", "@type": "owl:NamedIndividual",
         "label": "Sample Agent Query Patterns",
         "definition": "Worked examples mapping natural-language questions to ontology traversals and data product lookups.",
         "patterns": [
             {"question": "What is the current NAV of the fund?",
              "traversal": "CLOFund → hasPerformanceSnapshot → latest PerformanceSnapshot → nav",
              "dataProduct": "dp:DP-03",
              "metric": "clo:nav"},
             {"question": "Is the Class A OC test passing?",
              "traversal": "CLOFund → hasComplianceTest [OCTest where trancheClass='A'] → passFail, cushion",
              "dataProduct": "dp:DP-04",
              "metric": "clo:passFail, clo:cushion, clo:currentValue vs clo:threshold"},
             {"question": "What is the fund's current WARF?",
              "traversal": "CLOFund → hasKeyMetrics → latest KeyMetricSnapshot → warf",
              "dataProduct": "dp:DP-07",
              "metric": "clo:warf"},
             {"question": "Which obligors are in the CCC bucket?",
              "traversal": "CLOFund → hasAsset → LoanAsset where rating ≤ Caa1 → hasObligor",
              "dataProduct": "dp:DP-02 (filtered) + dp:DP-04 (threshold)",
              "metric": "Obligor name, par value, rating, % of pool"},
             {"question": "How much did equity receive in the last distribution?",
              "traversal": "CLOFund → hasCashflow [EquityDistribution, latest paymentDate] → amount",
              "dataProduct": "dp:DP-05",
              "metric": "clo:EquityDistribution amount (USD)"},
             {"question": "What happens if the Class B OC test breaches?",
              "traversal": "OCTest [class='B'] → breachConsequence + clo:WaterfallPriorityAxiom + clo:OCTestFormulaAxiom",
              "dataProduct": "Ontology axiom (no data lookup needed)",
              "metric": "Causal rule: cash diverted to senior principal repayment"},
             {"question": "What is the fund's net IRR since inception?",
              "traversal": "CLOFund → hasPerformanceSnapshot [latest] → netIRR",
              "dataProduct": "dp:DP-03",
              "metric": "clo:netIRR"},
             {"question": "Are we still in the reinvestment period?",
              "traversal": "CLOFund → reinvestmentPeriodEnd vs today + clo:ReinvestmentPeriodAxiom",
              "dataProduct": "dp:DP-01",
              "metric": "clo:reinvestmentPeriodEnd vs current date"},
             {"question": "What is total management fee paid this year?",
              "traversal": "CLOFund → hasFee [ManagementFee] → cumulativeAmountPaid [YTD filter]",
              "dataProduct": "dp:DP-06",
              "metric": "YTD management fee paid (USD)"},
         ]},

        # ---------- Fund instance binding (links ontology to actual workbook rows) ----------
        {"@id": f"clo:Fund/{FUND_ID}", "@type": "clo:CLOFund",
         "label": "DKIG Funding 2024-VII LLC",
         "fundName": "DKIG Funding 2024-VII LLC",
         "managerName": "DKIG Asset Management LLC",
         "vintageYear": 2024,
         "baseCurrency": "USD",
         "targetPar": 500000000,
         "reinvestmentPeriodEnd": "2029-03-15",
         "nonCallPeriodEnd": "2026-03-15",
         "legalMaturity": "2037-03-15",
         "asOfDate": REPORTING_DATE,
         "instanceData": {
             "staticProfile":      "DP-01 Static Profile",
             "portfolio":          "DP-02 Portfolio Snapshot",
             "performance":        "DP-03 Performance",
             "compliance":         "DP-04 Compliance",
             "cashflows":          "DP-05 Cashflows",
             "fees":               "DP-06 Fees",
             "keyMetrics":         "DP-07 Key Metrics",
             "liabilityStructure": "DP-08 Liability Structure",
         }},
    ],
}

with open(ONT_PATH, "w") as f:
    json.dump(ontology, f, indent=2)
print(f"[ok] wrote {ONT_PATH.name}")
